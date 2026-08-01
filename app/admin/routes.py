from flask import render_template, redirect, url_for, flash, request, jsonify, current_app, abort
from flask_login import login_required, current_user
from app import db
from app.admin import bp
from app.admin.forms import EmployeeForm, ServiceForm, VehicleSizeForm, CityForm, NeighborhoodForm, ProductForm, SubscriptionPackageForm, SiteSettingsForm, NotificationForm, AdminUserForm
from app.models import User, Service, VehicleSize, City, Neighborhood, Booking, Product, SubscriptionPackage, Subscription, EmployeeSchedule, SiteSettings, Notification, PushSubscription, BookingProduct, DiscountCode, Announcement, EmployeeLocation, CityServicePrice, CityProductPrice, PolishingOrder, Warehouse, CheckoutSession
from sqlalchemy import func, or_, extract
from datetime import date, timedelta, time, datetime
from werkzeug.utils import secure_filename
from urllib.parse import quote
import os
from pywebpush import webpush, WebPushException
import json
import secrets
import string
from app.utils.employee_breaks import employee_on_break_at, employee_break_overlaps
from app.utils.shift_utils import get_booking_work_date

ABANDONED_CHECKOUT_RETENTION_DAYS = 2

# Supervisors run day-to-day operations. Configuration, pricing, account
# administration, backups, and irreversible deletes remain admin-only.
ADMIN_ONLY_ENDPOINTS = {
    'admin.seasons', 'admin.add_season', 'admin.edit_season', 'admin.delete_season',
    'admin.services', 'admin.add_service', 'admin.edit_service', 'admin.delete_service',
    'admin.vehicle_sizes', 'admin.add_vehicle_size', 'admin.edit_vehicle_size',
    'admin.delete_vehicle_size', 'admin.locations', 'admin.add_city', 'admin.edit_city',
    'admin.delete_city', 'admin.add_neighborhood', 'admin.edit_neighborhood',
    'admin.delete_neighborhood', 'admin.packages', 'admin.add_package',
    'admin.edit_package', 'admin.delete_package', 'admin.assign_package_to_city',
    'admin.update_package_city_price', 'admin.remove_package_city_price',
    'admin.get_city_package_prices', 'admin.add_product', 'admin.edit_product',
    'admin.delete_product', 'admin.add_warehouse', 'admin.edit_warehouse',
    'admin.delete_warehouse', 'admin.assign_service_to_city_size',
    'admin.update_service_city_price', 'admin.remove_service_city_price',
    'admin.get_city_service_prices', 'admin.duplicate_service',
    'admin.assign_product_to_city', 'admin.update_product_city_price',
    'admin.remove_product_city_price', 'admin.get_city_product_prices',
    'admin.duplicate_product', 'admin.loyalty_settings', 'admin.backup_json',
    'admin.settings', 'admin.admins', 'admin.add_admin', 'admin.edit_admin',
    'admin.delete_admin', 'admin.send_notification', 'admin.announcements',
    'admin.add_announcement', 'admin.edit_announcement', 'admin.delete_announcement',
    'admin.toggle_announcement', 'admin.referral_tracking', 'admin.influencer_codes',
    'admin.add_influencer_code', 'admin.edit_influencer_code',
    'admin.toggle_influencer_code', 'admin.delete_influencer_code',
    'admin.delete_employee', 'admin.delete_customer', 'admin.delete_subscription',
    'admin.delete_polishing_order', 'admin.delete_booking_item', 'admin.delete_booking',
}

@bp.before_request
def before_request():
    if not current_user.is_authenticated or current_user.role not in ['admin', 'supervisor']:
        return redirect(url_for('auth.login'))
    if current_user.role == 'supervisor' and request.endpoint in ADMIN_ONLY_ENDPOINTS:
        abort(403)

def _supervisor_neighborhood_ids(user=None):
    user = user or current_user
    if not user.is_authenticated or user.role != 'supervisor':
        return None

    neighborhood_ids = set()
    for neighborhood in user.supervisor_neighborhoods:
        neighborhood_ids.add(neighborhood.id)
    for city in user.supervisor_cities:
        neighborhood_ids.update(n.id for n in city.neighborhoods)
    return list(neighborhood_ids)

def _apply_neighborhood_scope(query, model):
    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is None:
        return query
    if neighborhood_ids:
        return query.filter(model.neighborhood_id.in_(neighborhood_ids))
    return query.filter(model.id == -1)

def _scoped_employee_query(include_break=False):
    query = User.query.filter_by(role='employee')
    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is not None:
        if neighborhood_ids:
            query = query.join(User.neighborhoods).filter(Neighborhood.id.in_(neighborhood_ids)).distinct()
        else:
            query = query.filter(User.id == -1)
    if not include_break:
        query = query.filter(User.is_on_break == False)
    return query

def _is_employee_on_break(employee, target_date=None, target_time=None):
    return employee_on_break_at(employee, target_date, target_time)

def _employee_has_scope_access(employee):
    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is None:
        return True
    return bool(set(n.id for n in employee.neighborhoods).intersection(neighborhood_ids))

def _booking_has_scope_access(booking):
    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is None:
        return True
    return bool(booking and booking.neighborhood_id in neighborhood_ids)


def _checkout_has_scope_access(checkout):
    if current_user.role != 'supervisor':
        return True
    allowed_city_ids = {city.id for city in current_user.supervisor_cities}
    allowed_neighborhood_ids = set(_supervisor_neighborhood_ids() or [])
    return bool(
        checkout
        and (
            checkout.neighborhood_id in allowed_neighborhood_ids
            or checkout.city_id in allowed_city_ids
        )
    )


def _filter_bookings_by_work_date(bookings, from_date, to_date=None):
    """Filter bookings by the employee shift date, including overnight work."""
    to_date = to_date or from_date
    employee_ids = {booking.employee_id for booking in bookings if booking.employee_id}
    schedules_by_employee = {employee_id: [] for employee_id in employee_ids}
    if employee_ids:
        schedules = EmployeeSchedule.query.filter(
            EmployeeSchedule.employee_id.in_(employee_ids),
            EmployeeSchedule.is_active == True
        ).all()
        for schedule in schedules:
            schedules_by_employee.setdefault(schedule.employee_id, []).append(schedule)

    return [
        booking for booking in bookings
        if from_date <= get_booking_work_date(
            booking, schedules_by_employee.get(booking.employee_id, [])
        ) <= to_date
    ]

def _supervisor_city_ids(neighborhood_ids=None):
    neighborhood_ids = _supervisor_neighborhood_ids() if neighborhood_ids is None else neighborhood_ids
    if neighborhood_ids is None:
        return None
    city_ids = {city.id for city in current_user.supervisor_cities}
    if not neighborhood_ids:
        return list(city_ids)
    rows = Neighborhood.query.filter(Neighborhood.id.in_(neighborhood_ids)).all()
    city_ids.update(row.city_id for row in rows)
    return list(city_ids)

def _warehouse_has_scope_access(warehouse):
    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is None:
        return True
    if not warehouse or not warehouse.is_active:
        return False

    allowed_neighborhood_ids = set(neighborhood_ids)
    warehouse_neighborhood_ids = {n.id for n in warehouse.neighborhoods}
    if warehouse_neighborhood_ids:
        return bool(warehouse_neighborhood_ids.intersection(allowed_neighborhood_ids))

    allowed_city_ids = set(_supervisor_city_ids(neighborhood_ids))
    warehouse_city_ids = {c.id for c in warehouse.cities}
    return bool(warehouse_city_ids.intersection(allowed_city_ids))

def _normalize_whatsapp_phone(phone):
    """Return a digits-only international number suitable for wa.me links."""
    if not phone:
        return ''
    translated = str(phone).translate(str.maketrans(
        '٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹',
        '01234567890123456789'
    ))
    digits = ''.join(character for character in translated if character.isdigit())
    if digits.startswith('00'):
        digits = digits[2:]
    if digits.startswith('0'):
        digits = '966' + digits[1:]
    elif len(digits) == 9 and digits.startswith('5'):
        digits = '966' + digits
    return digits

def _scoped_warehouses():
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.id.desc()).all()
    if current_user.role == 'supervisor':
        warehouses = [warehouse for warehouse in warehouses if _warehouse_has_scope_access(warehouse)]
    return warehouses

@bp.route('/')
def index():
    from app.utils.timezone import get_saudi_date
    from datetime import datetime
    selected_date_str = request.args.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = get_saudi_date()
    else:
        selected_date = get_saudi_date()

    # Get supervisor's neighborhood scope if applicable
    supervisor_neighborhood_ids = []
    if current_user.role == 'supervisor':
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
    
    # Count employees (filter by scope for supervisors)
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            employees_count = User.query.filter_by(role='employee').join(User.neighborhoods).filter(Neighborhood.id.in_(supervisor_neighborhood_ids)).distinct().count()
            # Fix AmbiguousForeignKeysError by specifying join condition
            customers_count = User.query.filter_by(role='customer').join(Booking, User.id == Booking.customer_id).filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids)).distinct().count()
            dashboard_candidates = Booking.query.filter(
                Booking.neighborhood_id.in_(supervisor_neighborhood_ids),
                Booking.date.in_([selected_date, selected_date + timedelta(days=1)]),
                Booking.status != 'cancelled'
            ).all()
            bookings_count = len(_filter_bookings_by_work_date(dashboard_candidates, selected_date))
        else:
            employees_count = 0
            customers_count = 0
            bookings_count = 0
    else:
        employees_count = User.query.filter_by(role='employee').count()
        customers_count = User.query.filter_by(role='customer').count()
        dashboard_candidates = Booking.query.filter(
            Booking.date.in_([selected_date, selected_date + timedelta(days=1)]),
            Booking.status != 'cancelled'
        ).all()
        bookings_count = len(_filter_bookings_by_work_date(dashboard_candidates, selected_date))
    
    # Calculate total revenue from completed bookings
    completed_bookings_query = Booking.query.filter(
        Booking.status == 'completed',
        Booking.date.in_([selected_date, selected_date + timedelta(days=1)])
    )
    
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            completed_bookings_query = completed_bookings_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            completed_bookings_query = completed_bookings_query.filter_by(id=-1) # Empty result
            
    completed_bookings = _filter_bookings_by_work_date(
        completed_bookings_query.all(), selected_date
    )
    total_revenue = sum(b.service.price for b in completed_bookings if b.service)
    
    # Get recent bookings
    recent_bookings_query = Booking.query.filter(
        Booking.date.in_([selected_date, selected_date + timedelta(days=1)])
    ).order_by(Booking.date.desc(), Booking.time.desc())
    
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            recent_bookings_query = recent_bookings_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            recent_bookings_query = recent_bookings_query.filter_by(id=-1)
            
    recent_bookings = _filter_bookings_by_work_date(
        recent_bookings_query.all(), selected_date
    )[:5]
    
    return render_template('admin/index.html', 
                           employees_count=employees_count, 
                           customers_count=customers_count, 
                           bookings_count=bookings_count,
                           total_revenue=total_revenue,
                           recent_bookings=recent_bookings,
                           selected_date=selected_date)

# --- Seasons Management ---
@bp.route('/seasons')
def seasons():
    from app.models import Season, Booking
    seasons = Season.query.order_by(Season.start_date.desc()).all()
    
    # Calculate stats for each season
    for season in seasons:
        # Get all bookings within the season date range
        season_bookings = Booking.query.filter(
            Booking.date >= season.start_date,
            Booking.date <= season.end_date
        ).all()
        
        season.total_bookings = len(season_bookings)
        completed_bookings = [b for b in season_bookings if b.status == 'completed']
        season.completed_bookings = len(completed_bookings)
        
        # Calculate revenue for completed bookings
        total_revenue = 0
        for b in completed_bookings:
            # Calculate Service Revenue
            if b.subscription_id or b.used_free_wash:
                final_service_price = 0
            else:
                service_price = b.custom_service_price if b.custom_service_price is not None else (b.service.price if b.service else 0)
                vehicle_size_price = b.vehicle_size_price or 0
                discount_amount = 0
                
                if b.discount_code:
                    if b.discount_code.discount_type == 'percentage':
                        discount_amount = (service_price + vehicle_size_price) * (b.discount_code.value / 100)
                    else:
                        discount_amount = b.discount_code.value
                
                final_service_price = max(0, service_price + vehicle_size_price - discount_amount)
            
            # Calculate Products Revenue
            products_total = sum([(bp.unit_price if bp.unit_price is not None else bp.product.price) * bp.quantity for bp in b.products])
            
            total_revenue += (final_service_price + products_total)
            
        season.total_revenue = total_revenue

    return render_template('admin/seasons.html', seasons=seasons)

@bp.route('/seasons/add', methods=['GET', 'POST'])
def add_season():
    from app.admin.forms import SeasonForm
    from app.models import Season, Service, Product, SeasonalServicePrice, SeasonalProductPrice
    from datetime import datetime
    
    form = SeasonForm()
    services = Service.query.all()
    products = Product.query.all()
    
    if form.validate_on_submit():
        season = Season(
            name_ar=form.name_ar.data,
            name_en=form.name_en.data,
            start_date=datetime.strptime(form.start_date.data, '%Y-%m-%d').date(),
            end_date=datetime.strptime(form.end_date.data, '%Y-%m-%d').date(),
            is_active=form.is_active.data,
            allow_free_washes=form.allow_free_washes.data
        )
        db.session.add(season)
        db.session.flush() # Get season ID

        # Save service prices
        for service in services:
            price_val = request.form.get(f'service_price_{service.id}')
            if price_val and price_val.strip():
                try:
                    price = float(price_val)
                    ssp = SeasonalServicePrice(season_id=season.id, service_id=service.id, price=price)
                    db.session.add(ssp)
                except ValueError:
                    pass

        # Save product prices
        for product in products:
            price_val = request.form.get(f'product_price_{product.id}')
            if price_val and price_val.strip():
                try:
                    price = float(price_val)
                    spp = SeasonalProductPrice(season_id=season.id, product_id=product.id, price=price)
                    db.session.add(spp)
                except ValueError:
                    pass
                    
        db.session.commit()
        flash('تمت إضافة الموسم بنجاح', 'success')
        return redirect(url_for('admin.seasons'))
        
    return render_template('admin/season_form.html', form=form, services=services, products=products, ssp_dict={}, spp_dict={}, title='إضافة موسم')

@bp.route('/seasons/<int:id>/edit', methods=['GET', 'POST'])
def edit_season(id):
    from app.admin.forms import SeasonForm
    from app.models import Season, Service, Product, SeasonalServicePrice, SeasonalProductPrice
    from datetime import datetime
    
    season = Season.query.get_or_404(id)
    form = SeasonForm(obj=season)
    services = Service.query.all()
    products = Product.query.all()
    
    if request.method == 'POST' and form.validate_on_submit():
        season.name_ar = form.name_ar.data
        season.name_en = form.name_en.data
        season.start_date = datetime.strptime(form.start_date.data, '%Y-%m-%d').date()
        season.end_date = datetime.strptime(form.end_date.data, '%Y-%m-%d').date()
        season.is_active = form.is_active.data
        season.allow_free_washes = form.allow_free_washes.data
        
        # Clear old prices
        SeasonalServicePrice.query.filter_by(season_id=season.id).delete()
        SeasonalProductPrice.query.filter_by(season_id=season.id).delete()
        
        # Save new service prices
        for service in services:
            price_val = request.form.get(f'service_price_{service.id}')
            if price_val and price_val.strip():
                try:
                    price = float(price_val)
                    ssp = SeasonalServicePrice(season_id=season.id, service_id=service.id, price=price)
                    db.session.add(ssp)
                except ValueError:
                    pass

        # Save new product prices
        for product in products:
            price_val = request.form.get(f'product_price_{product.id}')
            if price_val and price_val.strip():
                try:
                    price = float(price_val)
                    spp = SeasonalProductPrice(season_id=season.id, product_id=product.id, price=price)
                    db.session.add(spp)
                except ValueError:
                    pass
                    
        db.session.commit()
        flash('تم تحديث الموسم بنجاح', 'success')
        return redirect(url_for('admin.seasons'))
        
    # GET: Populate dates
    if request.method == 'GET':
        form.start_date.data = season.start_date.strftime('%Y-%m-%d')
        form.end_date.data = season.end_date.strftime('%Y-%m-%d')
        
    # Fetch existing prices to populate template
    ssp_dict = {ssp.service_id: ssp.price for ssp in season.service_prices}
    spp_dict = {spp.product_id: spp.price for spp in season.product_prices}

    return render_template('admin/season_form.html', form=form, season=season, services=services, 
                           products=products, ssp_dict=ssp_dict, spp_dict=spp_dict, title='تعديل موسم')

@bp.route('/seasons/<int:id>/delete', methods=['POST'])
def delete_season(id):
    from app.models import Season
    season = Season.query.get_or_404(id)
    db.session.delete(season)
    db.session.commit()
    flash('تم حذف الموسم بنجاح', 'success')
    return redirect(url_for('admin.seasons'))

# --- Employee Management ---
@bp.route('/employees')
def employees():
    search_query = request.args.get('q', '').strip()
    role_filter = request.args.get('role', 'all')

    # Base query
    query = User.query

    if role_filter == 'employee':
        query = query.filter_by(role='employee')
    elif role_filter == 'supervisor':
        query = query.filter_by(role='supervisor')
    else:
        query = query.filter(User.role.in_(['employee', 'supervisor']))

    if current_user.role == 'supervisor':
        if role_filter == 'supervisor':
            query = query.filter(User.id == -1)
        else:
            neighborhood_ids = _supervisor_neighborhood_ids()
            if neighborhood_ids:
                query = query.join(User.neighborhoods).filter(Neighborhood.id.in_(neighborhood_ids)).distinct()
            else:
                query = query.filter(User.id == -1)

    if search_query:
        query = query.filter(
            or_(
                User.username.ilike(f"%{search_query}%"),
                User.phone.ilike(f"%{search_query}%"),
                User.email.ilike(f"%{search_query}%")
            )
        )
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(User.id.desc()).paginate(page=page, per_page=50, error_out=False)
    employees = pagination.items
    # Counters for tabs
    if current_user.role == 'supervisor':
        employee_count = _scoped_employee_query(include_break=True).count()
        all_count = employee_count
        supervisor_count = 0
    else:
        all_count = User.query.filter(User.role.in_(['employee', 'supervisor'])).count()
        employee_count = User.query.filter_by(role='employee').count()
        supervisor_count = User.query.filter_by(role='supervisor').count()

    return render_template(
        'admin/employees.html',
        employees=employees,
        role_filter=role_filter,
        all_count=all_count,
        employee_count=employee_count,
        supervisor_count=supervisor_count,
        pagination=pagination,
    )

@bp.route('/employees/add', methods=['GET', 'POST'])
def add_employee():
    from app.models import EmployeeSchedule
    from datetime import time
    
    form = EmployeeForm()
    
    # Populate choices
    all_neighborhoods = Neighborhood.query.join(City).all()
    all_cities = City.query.all()
    
    form.neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods]
    form.supervisor_cities.choices = [(c.id, c.name_ar) for c in all_cities]
    form.supervisor_neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods]
    
    # Restrict for supervisor (the current user, not the one being created)
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        # Filter choices to only show neighborhoods within supervisor's scope
        form.neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods if n.id in supervisor_neighborhood_ids]
    
    if form.validate_on_submit():
        # Check if phone number already exists
        existing_phone = User.query.filter_by(phone=form.phone.data).first()
        if existing_phone:
            flash('رقم الهاتف مستخدم بالفعل. الرجاء استخدام رقم هاتف آخر.', 'error')
            return render_template('admin/employee_form.html', form=form, title='إضافة موظف / مشرف')
        
        # Check if username already exists
        existing_username = User.query.filter_by(username=form.username.data).first()
        if existing_username:
            flash('اسم المستخدم موجود بالفعل. الرجاء اختيار اسم مستخدم آخر.', 'error')
            return render_template('admin/employee_form.html', form=form, title='إضافة موظف / مشرف')
        
        # Check if email already exists
        existing_email = User.query.filter_by(email=form.email.data).first()
        if existing_email:
            flash('البريد الإلكتروني مستخدم بالفعل. الرجاء استخدام بريد آخر.', 'error')
            return render_template('admin/employee_form.html', form=form, title='إضافة موظف / مشرف')
        
        # Determine role - only admin can create supervisors
        if current_user.role == 'admin' and form.role.data:
            role = form.role.data
        else:
            role = 'employee'
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            phone=form.phone.data,
            role=role,
            is_on_break=form.is_on_break.data if role == 'employee' else False
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.flush()  # Get user ID
        
        # Handle based on role
        if role == 'supervisor':
            # Assign supervisor cities
            city_ids = request.form.getlist('supervisor_cities')
            for city_id in city_ids:
                city = City.query.get(int(city_id))
                if city:
                    user.supervisor_cities.append(city)
            
            # Assign supervisor neighborhoods
            neighborhood_ids = request.form.getlist('supervisor_neighborhoods')
            for neighborhood_id in neighborhood_ids:
                neighborhood = Neighborhood.query.get(int(neighborhood_id))
                if neighborhood:
                    user.supervisor_neighborhoods.append(neighborhood)
        else:
            # Assign employee neighborhoods
            neighborhood_ids = request.form.getlist('neighborhoods')
            
            # Validate supervisor scope if current user is supervisor
            if current_user.role == 'supervisor':
                allowed_ids = set(supervisor_neighborhood_ids)
                neighborhood_ids = [nid for nid in neighborhood_ids if int(nid) in allowed_ids]
                
            for neighborhood_id in neighborhood_ids:
                neighborhood = Neighborhood.query.get(int(neighborhood_id))
                if neighborhood:
                    user.neighborhoods.append(neighborhood)
            
            # Create default schedule for employees only (Sun-Thu, 8 AM - 8 PM)
            for day in [6, 0, 1, 2, 3]:  # Sunday(6) to Thursday(3)
                schedule = EmployeeSchedule(
                    employee_id=user.id,
                    day_of_week=day,
                    start_time=time(8, 0),
                    end_time=time(20, 0),
                    is_active=True
                )
                db.session.add(schedule)
        
        db.session.commit()
        flash(f'تم إضافة {"المشرف" if role == "supervisor" else "الموظف"} بنجاح')
        return redirect(url_for('admin.employees'))
    return render_template('admin/employee_form.html', form=form, title='إضافة موظف / مشرف')

@bp.route('/employees/edit/<int:id>', methods=['GET', 'POST'])
def edit_employee(id):
    employee = User.query.get_or_404(id)
    
    # Check supervisor access (only for employees, not supervisors being edited)
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        # Check if employee belongs to any of supervisor's neighborhoods
        if employee.role == 'employee':
            employee_neighborhood_ids = [n.id for n in employee.neighborhoods]
            has_access = False
            
            for nid in employee_neighborhood_ids:
                if nid in supervisor_neighborhood_ids:
                    has_access = True
                    break
            
            if not has_access and employee_neighborhood_ids:
                flash('ليس لديك صلاحية لتعديل هذا الموظف', 'error')
                return redirect(url_for('admin.employees'))
        else:
            # Supervisors cannot edit other supervisors
            flash('ليس لديك صلاحية لتعديل المشرفين', 'error')
            return redirect(url_for('admin.employees'))

    form = EmployeeForm(obj=employee)
    
    all_neighborhoods = Neighborhood.query.join(City).all()
    all_cities = City.query.all()
    
    form.neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods]
    form.supervisor_cities.choices = [(c.id, c.name_ar) for c in all_cities]
    form.supervisor_neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods]
    
    # Restrict choices for supervisor (current user)
    if current_user.role == 'supervisor':
        form.neighborhoods.choices = [(n.id, f"{n.city.name_ar} - {n.name_ar}") for n in all_neighborhoods if n.id in supervisor_neighborhood_ids]

    if request.method == 'POST':
        # Update basic info
        employee.username = request.form.get('username')
        employee.email = request.form.get('email')
        employee.phone = request.form.get('phone')
        
        # Update role if admin is editing
        if current_user.role == 'admin' and form.role.data:
            old_role = employee.role
            new_role = form.role.data
            
            # If role changed, clear old associations
            if old_role != new_role:
                if old_role == 'employee':
                    employee.neighborhoods.clear()
                elif old_role == 'supervisor':
                    employee.supervisor_cities.clear()
                    employee.supervisor_neighborhoods.clear()
                
                employee.role = new_role
        
        # Update password if provided
        password = request.form.get('password')
        if password and password.strip():
            employee.set_password(password)

        employee.is_on_break = bool(request.form.get('is_on_break')) if employee.role == 'employee' else False
        
        # Handle based on role
        if employee.role == 'supervisor':
            # Update supervisor cities
            employee.supervisor_cities.clear()
            city_ids = request.form.getlist('supervisor_cities')
            for city_id in city_ids:
                city = City.query.get(int(city_id))
                if city:
                    employee.supervisor_cities.append(city)
            
            # Update supervisor neighborhoods
            employee.supervisor_neighborhoods.clear()
            neighborhood_ids = request.form.getlist('supervisor_neighborhoods')
            for neighborhood_id in neighborhood_ids:
                neighborhood = Neighborhood.query.get(int(neighborhood_id))
                if neighborhood:
                    employee.supervisor_neighborhoods.append(neighborhood)
        else:
            # Update employee neighborhoods
            neighborhood_ids = request.form.getlist('neighborhoods')
            
            if current_user.role == 'supervisor':
                # Get current neighborhoods outside scope (to preserve them)
                preserved_neighborhoods = [n for n in employee.neighborhoods if n.id not in supervisor_neighborhood_ids]
                
                # Filter new ids to be within scope
                new_scope_ids = [int(nid) for nid in neighborhood_ids if int(nid) in supervisor_neighborhood_ids]
                
                # Rebuild list
                employee.neighborhoods = preserved_neighborhoods
                for nid in new_scope_ids:
                    n = Neighborhood.query.get(nid)
                    if n:
                        employee.neighborhoods.append(n)
            else:
                # Admin: full replace
                employee.neighborhoods.clear()
                for neighborhood_id in neighborhood_ids:
                    neighborhood = Neighborhood.query.get(int(neighborhood_id))
                    if neighborhood and neighborhood not in employee.neighborhoods:
                        employee.neighborhoods.append(neighborhood)
        
        db.session.commit()
        flash('تم تعديل البيانات بنجاح')
        return redirect(url_for('admin.employees'))
    
    # GET request - pre-populate form
    form.username.data = employee.username
    form.email.data = employee.email
    form.phone.data = employee.phone
    form.is_on_break.data = bool(employee.is_on_break)
    
    if employee.role == 'supervisor':
        form.role.data = 'supervisor'
        form.supervisor_cities.data = [c.id for c in employee.supervisor_cities]
        form.supervisor_neighborhoods.data = [n.id for n in employee.supervisor_neighborhoods]
    else:
        form.role.data = 'employee'
        form.neighborhoods.data = [n.id for n in employee.neighborhoods]

    return render_template('admin/employee_form.html', form=form, title='تعديل موظف / مشرف', employee=employee)

@bp.route('/employees/<int:id>/toggle-break', methods=['POST'])
def toggle_employee_break(id):
    employee = User.query.get_or_404(id)

    if employee.role != 'employee':
        flash('خيار الراحة متاح للموظفين فقط', 'error')
        return redirect(request.referrer or url_for('admin.employees'))

    if current_user.role == 'supervisor' and not _employee_has_scope_access(employee):
        flash('Ù„ÙŠØ³ Ù„Ø¯ÙŠÙƒ ØµÙ„Ø§Ø­ÙŠØ© Ù„ØªØ¹Ø¯ÙŠÙ„ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆØ¸Ù', 'error')
        return redirect(request.referrer or url_for('admin.employees'))

    break_type = request.form.get('break_type')
    if employee.is_on_break and not break_type:
        employee.is_on_break = False
        employee.break_type = None
        employee.break_date = None
        employee.break_start_time = None
        employee.break_end_time = None
    else:
        employee.is_on_break = True
        employee.break_type = break_type if break_type in ['date', 'time', 'full_day'] else 'full_day'
        employee.break_date = None
        employee.break_start_time = None
        employee.break_end_time = None

        if employee.break_type == 'date':
            date_str = request.form.get('break_date')
            if date_str:
                employee.break_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                employee.break_type = 'full_day'
        elif employee.break_type == 'time':
            start_str = request.form.get('break_start_time')
            end_str = request.form.get('break_end_time')
            if start_str and end_str:
                employee.break_start_time = datetime.strptime(start_str, '%H:%M').time()
                employee.break_end_time = datetime.strptime(end_str, '%H:%M').time()
            else:
                employee.break_type = 'full_day'
    db.session.commit()
    flash('تم تفعيل الراحة للموظف' if employee.is_on_break else 'تم إيقاف الراحة للموظف', 'success')
    return redirect(request.referrer or url_for('admin.employees'))

@bp.route('/employees/schedule/<int:id>', methods=['GET', 'POST'])
def employee_schedule(id):
    employee = User.query.get_or_404(id)
    
    days_map = {
        'sunday': 6, 'monday': 0, 'tuesday': 1, 'wednesday': 2,
        'thursday': 3, 'friday': 4, 'saturday': 5
    }
    days_form = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    
    if request.method == 'POST':
        # Clear existing schedule
        EmployeeSchedule.query.filter_by(employee_id=id).delete()
        
        for day_name in days_form:
            enabled = request.form.get(f'{day_name}_enabled')
            if enabled:
                # Shift 1 (always required if day is enabled)
                start_time_str = request.form.get(f'{day_name}_start', '08:00')
                end_time_str = request.form.get(f'{day_name}_end', '20:00')
                
                try:
                    start_hour, start_min = map(int, start_time_str.split(':'))
                    end_hour, end_min = map(int, end_time_str.split(':'))
                    
                    schedule1 = EmployeeSchedule(
                        employee_id=id,
                        day_of_week=days_map[day_name],
                        shift_number=1,
                        start_time=time(start_hour, start_min),
                        end_time=time(end_hour, end_min),
                        is_active=True
                    )
                    db.session.add(schedule1)
                except ValueError:
                    pass
                
                # Shift 2 (optional)
                shift2_enabled = request.form.get(f'{day_name}_shift2_enabled')
                if shift2_enabled:
                    start2_time_str = request.form.get(f'{day_name}_start2', '')
                    end2_time_str = request.form.get(f'{day_name}_end2', '')
                    
                    if start2_time_str and end2_time_str:
                        try:
                            start2_hour, start2_min = map(int, start2_time_str.split(':'))
                            end2_hour, end2_min = map(int, end2_time_str.split(':'))
                            
                            schedule2 = EmployeeSchedule(
                                employee_id=id,
                                day_of_week=days_map[day_name],
                                shift_number=2,
                                start_time=time(start2_hour, start2_min),
                                end_time=time(end2_hour, end2_min),
                                is_active=True
                            )
                            db.session.add(schedule2)
                        except ValueError:
                            pass
        
        db.session.commit()
        flash('تم تحديث جدول العمل بنجاح')
        return redirect(url_for('admin.employees'))

    # GET: Prepare schedule data for template
    # Group schedules by day and shift
    schedules = {}
    for s in employee.schedules:
        if s.day_of_week not in schedules:
            schedules[s.day_of_week] = {}
        schedules[s.day_of_week][s.shift_number] = s
    
    return render_template('admin/employee_schedule.html', employee=employee, schedules=schedules)


@bp.route('/employees/delete/<int:id>', methods=['POST'])
def delete_employee(id):
    employee = User.query.get_or_404(id)
    db.session.delete(employee)
    db.session.commit()
    flash('تم حذف الموظف')
    return redirect(url_for('admin.employees'))

@bp.route('/employees/<int:id>/stats')
def employee_stats(id):
    employee = User.query.get_or_404(id)
    if current_user.role == 'supervisor' and (employee.role != 'employee' or not _employee_has_scope_access(employee)):
        flash('Ù„ÙŠØ³ Ù„Ø¯ÙŠÙƒ ØµÙ„Ø§Ø­ÙŠØ© Ù„Ø¹Ø±Ø¶ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆØ¸Ù', 'error')
        return redirect(url_for('admin.employees'))
    
    # Get date filters from query params
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    from_date = None
    to_date = None
    
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Load one additional calendar day so bookings after midnight can be
    # attributed to the previous overnight work shift.
    bookings_query = Booking.query.filter_by(employee_id=employee.id)
    if from_date:
        bookings_query = bookings_query.filter(Booking.date >= from_date)
    if to_date:
        bookings_query = bookings_query.filter(Booking.date <= to_date + timedelta(days=1))

    candidate_bookings = bookings_query.order_by(Booking.date.desc(), Booking.time.desc()).all()
    schedules = employee.schedules.filter_by(is_active=True).all()
    from app.utils.shift_utils import get_booking_work_date

    bookings = [
        booking for booking in candidate_bookings
        if (not from_date or get_booking_work_date(booking, schedules) >= from_date)
        and (not to_date or get_booking_work_date(booking, schedules) <= to_date)
    ]
    
    # Get all assigned subscriptions (with date filter if specified)
    subscriptions_query = Subscription.query.filter_by(employee_id=employee.id)
    if from_date:
        subscriptions_query = subscriptions_query.filter(Subscription.start_date >= from_date)
    if to_date:
        subscriptions_query = subscriptions_query.filter(Subscription.start_date <= to_date)
    subscriptions = subscriptions_query.all()
    
    # Get assigned neighborhoods
    neighborhoods = employee.neighborhoods
    
    # Calculate statistics (EXCLUDE cancelled from total count)
    non_cancelled_bookings = [b for b in bookings if b.status != 'cancelled']
    completed_bookings = [b for b in bookings if b.status == 'completed']
    active_subscriptions = len([s for s in subscriptions if s.status == 'active'])
    
    # Calculate earnings from completed bookings (EXCLUDE subscription bookings)
    total_earnings = 0
    cash_bookings_count = 0
    card_bookings_count = 0
    subscription_bookings_count = 0
    cash_earnings = 0
    card_earnings = 0
    
    for b in completed_bookings:
        # Track subscription count
        if b.subscription_id:
            subscription_bookings_count += 1
            
        # Calculate Service Revenue
        if b.subscription_id or b.used_free_wash:
            final_service_price = 0
        else:
            # Standard booking revenue
            service_price = b.custom_service_price if b.custom_service_price is not None else (b.service.price if b.service else 0)
            vehicle_size_price = b.vehicle_size_price or 0
            discount_amount = 0
            
            if b.discount_code:
                if b.discount_code.discount_type == 'percentage':
                    discount_amount = (service_price + vehicle_size_price) * (b.discount_code.value / 100)
                else:
                    discount_amount = b.discount_code.value
            
            final_service_price = max(0, service_price + vehicle_size_price - discount_amount)
        
        # Calculate Products Revenue (using custom unit_price if set)
        products_total = sum([(bp.unit_price if bp.unit_price is not None else bp.product.price) * bp.quantity for bp in b.products])
        
        booking_total = final_service_price + products_total
        
        total_earnings += booking_total
        
        # Count by payment method
        if b.payment_method == 'card':
            card_bookings_count += 1
            card_earnings += booking_total
        else:
            cash_bookings_count += 1
            cash_earnings += booking_total
    
    stats = {
        'total_bookings': len(non_cancelled_bookings),  # Exclude cancelled
        'cancelled_bookings': len([b for b in bookings if b.status == 'cancelled']),
        'completed_bookings': len(completed_bookings),
        'pending_bookings': len([b for b in bookings if b.status in ['pending', 'assigned', 'en_route', 'arrived', 'in_progress']]),
        'active_subscriptions': active_subscriptions,
        'total_subscriptions': len(subscriptions),
        'total_earnings': total_earnings,
        'cash_bookings_count': cash_bookings_count,
        'card_bookings_count': card_bookings_count,
        'subscription_bookings_count': subscription_bookings_count,
        'cash_earnings': cash_earnings,
        'card_earnings': card_earnings,
        'assigned_neighborhoods': len(neighborhoods)
    }
    
    # Format schedule for display
    days_map = {6: 'الأحد', 0: 'الاثنين', 1: 'الثلاثاء', 2: 'الأربعاء', 3: 'الخميس', 4: 'الجمعة', 5: 'السبت'}
    formatted_schedules = []
    for schedule in schedules:
        formatted_schedules.append({
            'day': days_map.get(schedule.day_of_week, ''),
            'start_time': schedule.start_time.strftime('%H:%M'),
            'end_time': schedule.end_time.strftime('%H:%M')
        })
    
    return render_template('admin/employee_stats.html',
                         employee=employee,
                         stats=stats,
                         bookings=bookings,
                         subscriptions=subscriptions,
                         neighborhoods=neighborhoods,
                         schedules=formatted_schedules,
                         from_date=from_date_str or '',
                         to_date=to_date_str or '')

# --- Customer Management ---
@bp.route('/customers')
def customers():
    search_query = request.args.get('q', '').strip()
    query = User.query.filter_by(role='customer')

    neighborhood_ids = _supervisor_neighborhood_ids()
    if neighborhood_ids is not None:
        if neighborhood_ids:
            query = query.join(Booking, User.id == Booking.customer_id).filter(Booking.neighborhood_id.in_(neighborhood_ids)).distinct()
        else:
            query = query.filter(User.id == -1)
    
    if search_query:
        query = query.filter(
            or_(
                User.username.ilike(f'%{search_query}%'),
                User.phone.ilike(f'%{search_query}%'),
                User.email.ilike(f'%{search_query}%')
            )
        )
        
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(User.id.desc()).paginate(page=page, per_page=50, error_out=False)
    customers = pagination.items
    
    # Calculate stats for each customer
    for customer in customers:
        # Count vehicles
        customer.vehicle_count = customer.vehicles.count()
        
        # Get last purchase date from bookings
        last_booking = Booking.query.filter_by(customer_id=customer.id).order_by(Booking.created_at.desc()).first()
        customer.last_purchase_date = last_booking.created_at if last_booking else None
        
    return render_template('admin/customers.html', customers=customers, pagination=pagination)

@bp.route('/customers/<int:id>/reset-password', methods=['POST'])
def reset_customer_password(id):
    customer = User.query.get_or_404(id)
    new_password = request.form.get('new_password', '123456')
    customer.set_password(new_password)
    db.session.commit()
    flash(f'تم تغيير كلمة السر للعميل {customer.username}')
    return redirect(url_for('admin.customers'))

@bp.route('/customers/<int:id>/add-points', methods=['POST'])
def add_points(id):
    customer = User.query.get_or_404(id)
    points = int(request.form.get('points', 0))
    customer.points = (customer.points or 0) + points
    db.session.commit()
    action = 'إضافة' if points >= 0 else 'خصم'
    flash(f'تم {action} {abs(points)} نقطة للعميل {customer.username}')
    return redirect(url_for('admin.customers'))

@bp.route('/customers/<int:id>/update-washes', methods=['POST'])
def update_washes(id):
    customer = User.query.get_or_404(id)
    washes = int(request.form.get('washes', 0))
    customer.free_washes = (customer.free_washes or 0) + washes
    db.session.commit()
    action = 'إضافة' if washes >= 0 else 'خصم'
    flash(f'تم {action} {abs(washes)} غسلة مجانية للعميل {customer.username}')
    return redirect(url_for('admin.customers'))

@bp.route('/customers/<int:id>/delete', methods=['POST'])
def delete_customer(id):
    """Delete customer and all related data"""
    from app.models import Vehicle, BookingProduct
    
    customer = User.query.get_or_404(id)
    
    # Verify customer role
    if customer.role != 'customer':
        flash('لا يمكن حذف هذا المستخدم', 'error')
        return redirect(url_for('admin.customers'))
    
    # Preserving Financial Data:
    # 1. Unlink Completed Bookings (keep them for reports)
    Booking.query.filter_by(customer_id=customer.id, status='completed').update({'customer_id': None})
    
    # 2. Unlink Active/Expired/Cancelled Subscriptions (if needed for history)
    # Ideally, we might want to keep expired ones for history too.
    # For now, let's keep 'completed' or 'expired' logic consistent if status exists
    # Assuming we want to keep all subscription history for revenue:
    Subscription.query.filter(Subscription.customer_id==customer.id).update({'customer_id': None})
    
    # Delete related data in correct order (only what wasn't preserved)
    
    # 3. Delete ONLY PENDING bookings (which are not completed)
    # Since we set customer_id=None for completed ones above, this loop will only find the remaining ones (pending, assigned, etc.)
    for booking in Booking.query.filter_by(customer_id=customer.id).all():
        BookingProduct.query.filter_by(booking_id=booking.id).delete()
        db.session.delete(booking)
    
    # 4. Update bookings where customer is employee (set to NULL) - rare edge case but good to keep
    Booking.query.filter_by(employee_id=customer.id).update({'employee_id': None})
    
    # 5. Delete vehicles
    Vehicle.query.filter_by(user_id=customer.id).delete()
    
    # 6. Delete notifications
    Notification.query.filter_by(user_id=customer.id).delete()
    
    # 7. Delete push subscriptions
    PushSubscription.query.filter_by(user_id=customer.id).delete()
    
    # Finally, delete the customer
    username = customer.username
    db.session.delete(customer)
    db.session.commit()
    
    flash(f'تم حذف العميل {username} بنجاح (مع الاحتفاظ بالسجلات المالية)', 'success')
    return redirect(url_for('admin.customers'))


@bp.route('/customers/<int:id>/ban', methods=['POST'])
def ban_customer(id):
    """Ban a customer permanently"""
    customer = User.query.get_or_404(id)
    
    if customer.role != 'customer':
        flash('لا يمكن حظر هذا المستخدم', 'error')
        return redirect(url_for('admin.customers'))
    
    ban_reason = request.form.get('ban_reason', '').strip()
    customer.is_banned = True
    customer.ban_reason = ban_reason if ban_reason else 'حظر بواسطة الإدارة'
    db.session.commit()
    
    flash(f'تم حظر العميل {customer.username} بنجاح', 'success')
    return redirect(url_for('admin.customers'))


@bp.route('/customers/<int:id>/unban', methods=['POST'])
def unban_customer(id):
    """Unban a customer"""
    customer = User.query.get_or_404(id)
    
    customer.is_banned = False
    customer.ban_reason = None
    db.session.commit()
    
    flash(f'تم إلغاء حظر العميل {customer.username} بنجاح', 'success')
    return redirect(url_for('admin.banned_customers'))


@bp.route('/banned-customers')
def banned_customers():
    """View all banned customers"""
    page = request.args.get('page', 1, type=int)
    pagination = User.query.filter_by(is_banned=True).order_by(User.id.desc()).paginate(page=page, per_page=50, error_out=False)
    banned = pagination.items
    return render_template('admin/banned_customers.html', customers=banned, pagination=pagination)

@bp.route('/customers/<int:id>/stats')
def customer_stats(id):
    customer = User.query.get_or_404(id)
    
    # Get all bookings for this customer
    bookings = Booking.query.filter_by(customer_id=customer.id).order_by(Booking.created_at.desc()).all()
    
    # Get all subscriptions for this customer  
    subscriptions = Subscription.query.filter_by(customer_id=customer.id).all()
    
    # Get all vehicles
    vehicles = customer.vehicles.all()
    
    # Calculate statistics
    total_bookings = len(bookings)
    completed_bookings = len([b for b in bookings if b.status == 'completed'])
    
    # Calculate accurate total spent including discounts and products
    total_spent = 0
    total_products_purchased = 0
    total_products_value = 0
    total_services_value = 0
    
    for booking in bookings:
        if booking.status == 'completed':
            # Calculate service price after discount/free wash
            service_price = booking.service.price if booking.service else 0
            discount_amount = 0
            
            # Check if free wash was used
            if booking.used_free_wash:
                service_price = 0
            # Check if discount code was applied
            elif booking.discount_code:
                if booking.discount_code.discount_type == 'percentage':
                    discount_amount = service_price * (booking.discount_code.value / 100)
                else:
                    discount_amount = booking.discount_code.value
            
            # Calculate final service price
            final_service_price = service_price - discount_amount
            total_services_value += final_service_price
            
            # Calculate products total
            # Calculate products total
            products_total = sum([(bp.unit_price if bp.unit_price is not None else bp.product.price) * bp.quantity for bp in booking.products])
            total_products_purchased += sum([bp.quantity for bp in booking.products])
            total_products_value += products_total
            
            # Add to total spent
            total_spent += final_service_price + products_total
    
    stats = {
        'total_bookings': total_bookings,
        'completed_bookings': completed_bookings,
        'pending_bookings': len([b for b in bookings if b.status in ['pending', 'assigned', 'en_route', 'in_progress']]),
        'total_spent': total_spent,
        'total_services_value': total_services_value,
        'total_products_purchased': total_products_purchased,
        'total_products_value': total_products_value,
        'points': customer.points or 0,
        'free_washes': customer.free_washes or 0,
        'total_vehicles': len(vehicles)
    }
    
    return render_template('admin/customer_stats.html', 
                         customer=customer, 
                         stats=stats, 
                         bookings=bookings, 
                         vehicles=vehicles)

@bp.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
def edit_customer(id):
    customer = User.query.get_or_404(id)
    
    if request.method == 'POST':
        # Update customer information
        customer.username = request.form.get('username')
        customer.email = request.form.get('email')
        customer.phone = request.form.get('phone')
        
        # Update password if provided
        new_password = request.form.get('password')
        if new_password and new_password.strip():
            customer.set_password(new_password)
        
        db.session.commit()
        flash(f'تم تحديث معلومات العميل {customer.username} بنجاح!')
        return redirect(url_for('admin.customers'))
    
    return render_template('admin/edit_customer.html', customer=customer)

@bp.route('/ratings')
def ratings():
    from datetime import datetime
    
    # Filters
    employee_id = request.args.get('employee_id')
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    query = Booking.query.filter(Booking.rating.isnot(None)).order_by(Booking.rating_date.desc())
    employees_query = _scoped_employee_query(include_break=True)
    scoped_employee_ids = None

    if current_user.role == 'supervisor':
        scoped_employee_ids = [employee.id for employee in employees_query.all()]
        query = query.filter(Booking.employee_id.in_(scoped_employee_ids)) if scoped_employee_ids else query.filter(Booking.id == -1)
    
    if employee_id:
        employee_id_int = int(employee_id)
        if scoped_employee_ids is not None and employee_id_int not in scoped_employee_ids:
            query = query.filter(Booking.id == -1)
        else:
            query = query.filter(Booking.employee_id == employee_id_int)
    
    # Date range filter
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Booking.rating_date) >= from_date)
        except ValueError:
            pass
            
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Booking.rating_date) <= to_date)
        except ValueError:
            pass
            
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(Booking.rating_date.desc() if hasattr(Booking, 'rating_date') else Booking.id.desc()).paginate(page=page, per_page=50, error_out=False)
    ratings = pagination.items
    employees = employees_query.all()
    
    return render_template('admin/ratings.html', ratings=ratings, employees=employees, pagination=pagination)

@bp.route('/customers/export')
def export_customers():
    try:
        import io
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        customers = User.query.filter_by(role='customer').all()
        
        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "العملاء"
        ws.sheet_view.rightToLeft = True  # Enable RTL
        
        # Headers
        headers = ['#', 'الاسم', 'رقم الجوال', 'البريد الإلكتروني', 'الغسلات المجانية', 'نقاط الولاء', 'عدد السيارات', 'تاريخ اخر عملية شراء']
        ws.append(headers)
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4DA8DA", end_color="4DA8DA", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply style to headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            
        # Data
        for customer in customers:
            # Count vehicles
            vehicle_count = customer.vehicles.count()
            
            # Get last purchase date - with fallback
            try:
                last_booking = Booking.query.filter_by(customer_id=customer.id).order_by(Booking.date.desc()).first()
                if last_booking:
                    # Try created_at first, fallback to date
                    if hasattr(last_booking, 'created_at') and last_booking.created_at:
                        last_purchase = last_booking.created_at.strftime('%Y-%m-%d')
                    else:
                        last_purchase = last_booking.date.strftime('%Y-%m-%d') if last_booking.date else '-'
                else:
                    last_purchase = '-'
            except Exception as e:
                last_purchase = '-'
            
            row = [
                customer.id,
                customer.username,
                customer.phone or '-',
                customer.email or '-',
                customer.free_washes or 0,
                customer.points or 0,
                vehicle_count,
                last_purchase
            ]
            ws.append(row)
            
            # Apply style to data rows
            for cell in ws[ws.max_row]:
                cell.alignment = alignment
                cell.border = border

        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='customers.xlsx'
        )
    except ImportError:
        # openpyxl not installed - fallback to CSV
        flash('خطأ: مكتبة openpyxl غير مثبتة. يرجى تثبيتها أولاً.', 'error')
        return redirect(url_for('admin.customers'))
    except Exception as e:
        # Log the error and show user-friendly message
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'error')
        return redirect(url_for('admin.customers'))


# --- Service Management ---
@bp.route('/services')
def services():
    page = request.args.get('page', 1, type=int)
    pagination = Service.query.order_by(Service.id.asc()).paginate(page=page, per_page=50, error_out=False)
    services_list = pagination.items
    services_data = []
    
    for service in services_list:
        # Calculate completed bookings count
        completed_bookings_count = Booking.query.filter_by(service_id=service.id, status='completed').count()
        
        # Calculate total revenue
        total_revenue = completed_bookings_count * service.price
        
        services_data.append({
            'service': service,
            'completed_bookings_count': completed_bookings_count,
            'total_revenue': total_revenue
        })
        
    cities = City.query.filter_by(is_active=True).all()
    vehicle_sizes = VehicleSize.query.filter_by(is_active=True).all()
    return render_template('admin/services.html', services=services_data, cities=cities, vehicle_sizes=vehicle_sizes, pagination=pagination)

@bp.route('/services/add', methods=['GET', 'POST'])
def add_service():
    form = ServiceForm()
    if form.validate_on_submit():
        try:
            service = Service(name_ar=form.name_ar.data, name_en=form.name_en.data, 
                              price=form.price.data, duration=form.duration.data, 
                              description=form.description.data,
                              includes_free_wash=form.includes_free_wash.data,
                              awards_loyalty_point=form.awards_loyalty_point.data,
                              is_active=form.is_active.data)
            db.session.add(service)
            db.session.commit()
            flash('تم إضافة الخدمة بنجاح', 'success')
            return redirect(url_for('admin.services'))
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
    elif request.method == 'POST':
        # Form validation failed
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    return render_template('admin/service_form.html', form=form, title='إضافة خدمة')

@bp.route('/services/edit/<int:id>', methods=['GET', 'POST'])
def edit_service(id):
    service = Service.query.get_or_404(id)
    form = ServiceForm(obj=service)
    if form.validate_on_submit():
        form.populate_obj(service)
        db.session.commit()
        flash('تم تعديل الخدمة')
        return redirect(url_for('admin.services'))
    return render_template('admin/service_form.html', form=form, title='تعديل خدمة')

@bp.route('/services/delete/<int:id>', methods=['POST'])
def delete_service(id):
    service = Service.query.get_or_404(id)
    db.session.delete(service)
    db.session.commit()
    flash('تم حذف الخدمة')
    return redirect(url_for('admin.services'))


# --- Vehicle Size Management ---
@bp.route('/vehicle-sizes')
def vehicle_sizes():
    sizes = VehicleSize.query.all()
    return render_template('admin/vehicle_sizes.html', vehicle_sizes=sizes)

@bp.route('/vehicle-sizes/add', methods=['GET', 'POST'])
def add_vehicle_size():
    form = VehicleSizeForm()
    if form.validate_on_submit():
        size = VehicleSize(
            name_ar=form.name_ar.data,
            name_en=form.name_en.data,
            price_adjustment=form.price_adjustment.data,
            is_active=form.is_active.data
        )
        db.session.add(size)
        db.session.commit()
        flash('تم إضافة حجم السيارة بنجاح')
        return redirect(url_for('admin.vehicle_sizes'))
    return render_template('admin/vehicle_size_form.html', form=form, title='إضافة حجم سيارة')

@bp.route('/vehicle-sizes/edit/<int:id>', methods=['GET', 'POST'])
def edit_vehicle_size(id):
    size = VehicleSize.query.get_or_404(id)
    form = VehicleSizeForm(obj=size)
    if form.validate_on_submit():
        size.name_ar = form.name_ar.data
        size.name_en = form.name_en.data
        size.price_adjustment = form.price_adjustment.data
        size.is_active = form.is_active.data
        db.session.commit()
        flash('تم تعديل حجم السيارة بنجاح')
        return redirect(url_for('admin.vehicle_sizes'))
    return render_template('admin/vehicle_size_form.html', form=form, title='تعديل حجم سيارة')

@bp.route('/vehicle-sizes/delete/<int:id>', methods=['POST'])
def delete_vehicle_size(id):
    size = VehicleSize.query.get_or_404(id)
    db.session.delete(size)
    db.session.commit()
    flash('تم حذف حجم السيارة')
    return redirect(url_for('admin.vehicle_sizes'))

# --- Products Management ---
@bp.route('/products')
def products():
    from app.models import ProductStock
    page = request.args.get('page', 1, type=int)
    pagination = Product.query.order_by(Product.id.asc()).paginate(page=page, per_page=50, error_out=False)
    all_products = pagination.items
    products_data = []
    supervisor_neighborhood_ids_for_products = _supervisor_neighborhood_ids()
    allowed_warehouse_ids_for_products = None
    if current_user.role == 'supervisor':
        allowed_warehouse_ids_for_products = [warehouse.id for warehouse in _scoped_warehouses()]
    
    # Calculate global total revenue across ALL products
    total_sales_revenue_query = db.session.query(
        func.sum(
            BookingProduct.quantity * func.coalesce(BookingProduct.unit_price, Product.price)
        )
    ).join(Booking, BookingProduct.booking_id == Booking.id)\
     .join(Product, BookingProduct.product_id == Product.id)\
     .filter(Booking.status == 'completed')
    if supervisor_neighborhood_ids_for_products is not None:
        total_sales_revenue_query = total_sales_revenue_query.filter(
            Booking.neighborhood_id.in_(supervisor_neighborhood_ids_for_products)
        ) if supervisor_neighborhood_ids_for_products else total_sales_revenue_query.filter(Booking.id == -1)
    total_sales_revenue = total_sales_revenue_query.scalar() or 0
    
    for product in all_products:
        # Get total quantity sold (only for completed bookings)
        sold_quantity_query = db.session.query(func.sum(BookingProduct.quantity))\
            .join(Booking, BookingProduct.booking_id == Booking.id)\
            .filter(
                BookingProduct.product_id == product.id,
                Booking.status == 'completed'
            )
        if supervisor_neighborhood_ids_for_products is not None:
            sold_quantity_query = sold_quantity_query.filter(
                Booking.neighborhood_id.in_(supervisor_neighborhood_ids_for_products)
            ) if supervisor_neighborhood_ids_for_products else sold_quantity_query.filter(Booking.id == -1)
        sold_quantity = sold_quantity_query.scalar() or 0
        
        # Calculate revenue using unit_price if available, else product.price
        revenue_query = db.session.query(
            func.sum(
                BookingProduct.quantity * func.coalesce(BookingProduct.unit_price, Product.price)
            )
        ).join(Booking, BookingProduct.booking_id == Booking.id)\
         .join(Product, BookingProduct.product_id == Product.id)\
         .filter(
            BookingProduct.product_id == product.id,
            Booking.status == 'completed'
        )
        if supervisor_neighborhood_ids_for_products is not None:
            revenue_query = revenue_query.filter(
                Booking.neighborhood_id.in_(supervisor_neighborhood_ids_for_products)
            ) if supervisor_neighborhood_ids_for_products else revenue_query.filter(Booking.id == -1)
        revenue = revenue_query.scalar() or 0
        
        warehouse_stock_query = db.session.query(func.sum(ProductStock.quantity)).filter(
            ProductStock.product_id == product.id,
            ProductStock.warehouse_id.isnot(None)
        )
        if allowed_warehouse_ids_for_products is not None:
            warehouse_stock_query = warehouse_stock_query.filter(
                ProductStock.warehouse_id.in_(allowed_warehouse_ids_for_products)
            ) if allowed_warehouse_ids_for_products else warehouse_stock_query.filter(ProductStock.id == -1)
        warehouse_stock = warehouse_stock_query.scalar()
        display_stock = warehouse_stock if warehouse_stock is not None else (0 if current_user.role == 'supervisor' else product.stock_quantity)
        # total_sales_revenue is now calculated globally above the loop

        
        products_data.append({
            'product': product,
            'sold_quantity': sold_quantity,
            'revenue': revenue,
            'stock': display_stock
        })
    
    # Get cities for location stock management
    # For supervisors, only show their assigned areas
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        supervisor_city_ids = set()
        
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
            for n in current_user.supervisor_neighborhoods:
                supervisor_city_ids.add(n.city_id)
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_city_ids.add(city.id)
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        cities = City.query.filter(City.id.in_(supervisor_city_ids), City.is_active==True).all()
        # Only include neighborhoods in supervisor's scope
        cities_json = json.dumps([{
            'id': c.id,
            'name_ar': c.name_ar,
            'neighborhoods': [{'id': n.id, 'name_ar': n.name_ar} for n in c.neighborhoods if n.id in supervisor_neighborhood_ids]
        } for c in cities])
    else:
        cities = City.query.filter_by(is_active=True).all()
        cities_json = json.dumps([{
            'id': c.id,
            'name_ar': c.name_ar,
            'neighborhoods': [{'id': n.id, 'name_ar': n.name_ar} for n in c.neighborhoods]
        } for c in cities])
    
    warehouses = _scoped_warehouses()
    all_neighborhoods_query = Neighborhood.query.join(City).filter(Neighborhood.is_active == True)
    supervisor_neighborhood_ids = _supervisor_neighborhood_ids()
    if supervisor_neighborhood_ids is not None:
        all_neighborhoods_query = all_neighborhoods_query.filter(Neighborhood.id.in_(supervisor_neighborhood_ids)) if supervisor_neighborhood_ids else all_neighborhoods_query.filter(Neighborhood.id == -1)
    all_neighborhoods = all_neighborhoods_query.order_by(City.name_ar, Neighborhood.name_ar).all()
    return render_template('admin/products.html', products=products_data, 
                           total_sales_revenue=total_sales_revenue, cities=cities, cities_json=cities_json,
                           warehouses=warehouses, all_neighborhoods=all_neighborhoods, pagination=pagination)

@bp.route('/products/update_stock/<int:product_id>', methods=['POST'])
def update_stock(product_id):
    """Update product stock - either global or per location"""
    from app.models import ProductStock
    
    product = Product.query.get_or_404(product_id)
    stock = request.form.get('stock', 0, type=int)
    city_id = request.form.get('city_id', type=int)
    neighborhood_id = request.form.get('neighborhood_id', type=int)
    warehouse_id = request.form.get('warehouse_id', type=int)
    price = request.form.get('price', type=float)
    
    if warehouse_id:
        warehouse = Warehouse.query.get(warehouse_id)
        if not _warehouse_has_scope_access(warehouse):
            flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062a\u0639\u062f\u064a\u0644 \u0647\u0630\u0627 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639', 'error')
            return redirect(url_for('admin.products'))

        existing = ProductStock.query.filter_by(
            product_id=product_id,
            warehouse_id=warehouse_id
        ).first()
        
        if existing:
            existing.quantity = stock
            existing.price = price
        else:
            new_stock = ProductStock(
                product_id=product_id,
                warehouse_id=warehouse_id,
                quantity=stock,
                price=price
            )
            db.session.add(new_stock)
        
        flash('\u062a\u0645 \u062a\u062d\u062f\u064a\u062b \u0645\u062e\u0632\u0648\u0646 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639')
    elif city_id:
        # Legacy location-based stock update
        existing = ProductStock.query.filter_by(
            product_id=product_id,
            city_id=city_id,
            neighborhood_id=neighborhood_id if neighborhood_id else None
        ).first()
        
        if existing:
            existing.quantity = stock
        else:
            new_stock = ProductStock(
                product_id=product_id,
                city_id=city_id,
                neighborhood_id=neighborhood_id if neighborhood_id else None,
                quantity=stock
            )
            db.session.add(new_stock)
        
        flash(f'تم تحديث المخزون للموقع المحدد')
    else:
        # Global stock update
        product.stock_quantity = stock
        flash('تم تحديث المخزون العام')
    
    db.session.commit()
    return redirect(url_for('admin.products'))

@bp.route('/products/location_stock/<int:product_id>')
def get_location_stock(product_id):
    """Get all location stocks for a product"""
    from app.models import ProductStock
    
    stocks = ProductStock.query.filter_by(product_id=product_id).all()
    if current_user.role == 'supervisor':
        allowed_warehouse_ids = {warehouse.id for warehouse in _scoped_warehouses()}
        stocks = [stock for stock in stocks if stock.warehouse_id in allowed_warehouse_ids]
    result = []
    
    for s in stocks:
        result.append({
            'id': s.id,
            'city_id': s.city_id,
            'city_name': s.city.name_ar if s.city else '',
            'neighborhood_id': s.neighborhood_id,
            'neighborhood_name': s.neighborhood.name_ar if s.neighborhood else 'كل الأحياء',
            'warehouse_id': s.warehouse_id,
            'warehouse_name': s.warehouse.name_ar if s.warehouse else '',
            'quantity': s.quantity,
            'price': s.price
        })
    
    return jsonify(result)

@bp.route('/warehouses/add', methods=['POST'])
def add_warehouse():
    name_ar = request.form.get('name_ar', '').strip()
    name_en = request.form.get('name_en', '').strip()
    city_ids = request.form.getlist('city_ids')
    neighborhood_ids = request.form.getlist('neighborhood_ids')

    if not name_ar:
        flash('\u0627\u0633\u0645 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0645\u0637\u0644\u0648\u0628', 'error')
        return redirect(url_for('admin.products'))

    if not name_ar:
        flash('Ø§Ø³Ù… Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹ Ù…Ø·Ù„ÙˆØ¨', 'error')
        return redirect(url_for('admin.products'))

    supervisor_neighborhood_ids = _supervisor_neighborhood_ids()
    if supervisor_neighborhood_ids is not None:
        allowed_neighborhood_ids = set(supervisor_neighborhood_ids)
        allowed_city_ids = set(_supervisor_city_ids(supervisor_neighborhood_ids))
        city_ids = [cid for cid in city_ids if int(cid) in allowed_city_ids]
        neighborhood_ids = [nid for nid in neighborhood_ids if int(nid) in allowed_neighborhood_ids]
        if not city_ids and not neighborhood_ids:
            flash('\u064a\u062c\u0628 \u0631\u0628\u0637 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0628\u0645\u062f\u064a\u0646\u0629 \u0623\u0648 \u062d\u064a \u0636\u0645\u0646 \u0646\u0637\u0627\u0642\u0643', 'error')
            return redirect(url_for('admin.products'))

    warehouse = Warehouse(name_ar=name_ar, name_en=name_en or None, is_active=True)
    for city_id in city_ids:
        city = City.query.get(int(city_id))
        if city:
            warehouse.cities.append(city)
    for neighborhood_id in neighborhood_ids:
        neighborhood = Neighborhood.query.get(int(neighborhood_id))
        if neighborhood:
            warehouse.neighborhoods.append(neighborhood)

    db.session.add(warehouse)
    db.session.commit()
    flash('\u062a\u0645\u062a \u0625\u0636\u0627\u0641\u0629 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0628\u0646\u062c\u0627\u062d', 'success')
    return redirect(url_for('admin.products'))
    db.session.commit()
    flash('ØªÙ…Øª Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹ Ø¨Ù†Ø¬Ø§Ø­', 'success')
    return redirect(url_for('admin.products'))

@bp.route('/warehouses/<int:id>/edit', methods=['POST'])
def edit_warehouse(id):
    warehouse = Warehouse.query.get_or_404(id)
    if not _warehouse_has_scope_access(warehouse):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062a\u0639\u062f\u064a\u0644 \u0647\u0630\u0627 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639', 'error')
        return redirect(url_for('admin.products'))

    name_ar = request.form.get('name_ar', '').strip()
    name_en = request.form.get('name_en', '').strip()
    city_ids = request.form.getlist('city_ids')
    neighborhood_ids = request.form.getlist('neighborhood_ids')

    if not name_ar:
        flash('\u0627\u0633\u0645 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0645\u0637\u0644\u0648\u0628', 'error')
        return redirect(url_for('admin.products'))

    supervisor_neighborhood_ids = _supervisor_neighborhood_ids()
    if supervisor_neighborhood_ids is not None:
        allowed_neighborhood_ids = set(supervisor_neighborhood_ids)
        allowed_city_ids = set(_supervisor_city_ids(supervisor_neighborhood_ids))
        city_ids = [cid for cid in city_ids if int(cid) in allowed_city_ids]
        neighborhood_ids = [nid for nid in neighborhood_ids if int(nid) in allowed_neighborhood_ids]
        if not city_ids and not neighborhood_ids:
            flash('\u064a\u062c\u0628 \u0631\u0628\u0637 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0628\u0645\u062f\u064a\u0646\u0629 \u0623\u0648 \u062d\u064a \u0636\u0645\u0646 \u0646\u0637\u0627\u0642\u0643', 'error')
            return redirect(url_for('admin.products'))

    warehouse.name_ar = name_ar
    warehouse.name_en = name_en or None
    warehouse.cities = []
    warehouse.neighborhoods = []

    for city_id in city_ids:
        city = City.query.get(int(city_id))
        if city:
            warehouse.cities.append(city)
    for neighborhood_id in neighborhood_ids:
        neighborhood = Neighborhood.query.get(int(neighborhood_id))
        if neighborhood:
            warehouse.neighborhoods.append(neighborhood)

    db.session.commit()
    flash('\u062a\u0645 \u062a\u0639\u062f\u064a\u0644 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0628\u0646\u062c\u0627\u062d', 'success')
    return redirect(url_for('admin.products'))

@bp.route('/warehouses/<int:id>/delete', methods=['POST'])
def delete_warehouse(id):
    from app.models import ProductStock

    warehouse = Warehouse.query.get_or_404(id)
    if not _warehouse_has_scope_access(warehouse):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062d\u0630\u0641 \u0647\u0630\u0627 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639', 'error')
        return redirect(url_for('admin.products'))

    ProductStock.query.filter_by(warehouse_id=warehouse.id).delete(synchronize_session=False)
    warehouse.cities = []
    warehouse.neighborhoods = []
    db.session.delete(warehouse)
    db.session.commit()
    flash('\u062a\u0645 \u062d\u0630\u0641 \u0627\u0644\u0645\u0633\u062a\u0648\u062f\u0639 \u0648\u0628\u064a\u0627\u0646\u0627\u062a\u0647', 'success')
    return redirect(url_for('admin.products'))
    db.session.commit()
    flash('ØªÙ… Ø¥ÙŠÙ‚Ø§Ù Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹', 'success')
    return redirect(url_for('admin.products'))

@bp.route('/products/add', methods=['GET', 'POST'])
def add_product():
    import os
    from werkzeug.utils import secure_filename
    
    form = ProductForm()
    if form.validate_on_submit():
        image_url = None
        if form.image.data:
            file = form.image.data
            filename = secure_filename(file.filename)
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join('app', 'static', 'uploads', filename)
            file.save(filepath)
            image_url = f"/static/uploads/{filename}"
        
        product = Product(
            name_ar=form.name_ar.data,
            name_en=form.name_en.data,
            price=float(form.price.data),
            image_url=image_url
        )
        db.session.add(product)
        db.session.commit()
        flash('تم إضافة المنتج بنجاح')
        return redirect(url_for('admin.products'))
    return render_template('admin/product_form.html', form=form, title='إضافة منتج')


# --- Location Management ---
@bp.route('/locations')
def locations():
    cities = City.query.all()

    city_stats = []
    for city in cities:
        neighborhood_count = city.neighborhoods.count()

        city_neighborhood_ids = [n.id for n in city.neighborhoods]

        if city_neighborhood_ids:
            bookings_q = Booking.query.filter(Booking.neighborhood_id.in_(city_neighborhood_ids))
            booking_count = bookings_q.count()
            completed_bookings = bookings_q.filter_by(status='completed').all()
        else:
            booking_count = 0
            completed_bookings = []

        revenue = sum(b.service.price for b in completed_bookings if b.service)

        city_stats.append({
            'city': city,
            'neighborhood_count': neighborhood_count,
            'booking_count': booking_count,
            'revenue': revenue,
        })

    return render_template('admin/locations.html', city_stats=city_stats)

@bp.route('/locations/city/add', methods=['GET', 'POST'])
def add_city():
    if request.method == 'POST':
        # Handle JSON from OSM modal
        name_ar = request.form.get('name_ar', '').strip()
        name_en = request.form.get('name_en', '').strip()
        osm_place_id = request.form.get('osm_place_id', '').strip()
        
        if not name_ar or not name_en:
            flash('يجب إدخال اسم المدينة بالعربية والإنجليزية', 'error')
            return redirect(url_for('admin.locations'))
        
        # Check if city already exists
        existing = City.query.filter(
            (City.name_en == name_en) | (City.osm_place_id == osm_place_id)
        ).first() if osm_place_id else City.query.filter_by(name_en=name_en).first()
        
        if existing:
            flash('هذه المدينة موجودة بالفعل', 'error')
            return redirect(url_for('admin.locations'))
        
        city = City(
            name_ar=name_ar,
            name_en=name_en,
            osm_place_id=osm_place_id if osm_place_id else None,
            is_active=True
        )
        db.session.add(city)
        db.session.commit()
        flash('تم إضافة المدينة بنجاح', 'success')
        return redirect(url_for('admin.locations'))
    
    # GET: show simple form (fallback)
    form = CityForm()
    return render_template('admin/location_form.html', form=form, title='إضافة مدينة', type='city')

@bp.route('/locations/city/edit/<int:id>', methods=['GET', 'POST'])
def edit_city(id):
    city = City.query.get_or_404(id)
    form = CityForm(obj=city)
    if form.validate_on_submit():
        form.populate_obj(city)
        db.session.commit()
        flash('تم تعديل المدينة')
        return redirect(url_for('admin.locations'))
    return render_template('admin/location_form.html', form=form, title='تعديل مدينة', type='city')

@bp.route('/products/edit/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
    import os
    from werkzeug.utils import secure_filename
    
    product = Product.query.get_or_404(id)
    form = ProductForm()
    
    if form.validate_on_submit():
        product.name_ar = form.name_ar.data
        product.name_en = form.name_en.data
        product.price = float(form.price.data)
        
        # Handle image upload
        if form.image.data:
            file = form.image.data
            filename = secure_filename(file.filename)
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join('app', 'static', 'uploads', filename)
            file.save(filepath)
            product.image_url = f"/static/uploads/{filename}"
        
        db.session.commit()
        flash('تم تحديث المنتج')
        return redirect(url_for('admin.products'))
    elif request.method == 'GET':
        form.name_ar.data = product.name_ar
        form.name_en.data = product.name_en
        form.price.data = str(product.price)
    
    return render_template('admin/product_form.html', form=form, title='تعديل منتج', product=product)


@bp.route('/products/delete/<int:id>', methods=['POST'])
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج')
    return redirect(url_for('admin.products'))

@bp.route('/products/stats/<int:id>')
def product_stats(id):
    from sqlalchemy import func
    from app.models import ProductStock

    product = Product.query.get_or_404(id)
    supervisor_neighborhood_ids = _supervisor_neighborhood_ids()
    
    # Calculate total sold quantity
    sold_quantity_query = db.session.query(func.sum(BookingProduct.quantity))\
        .join(Booking, BookingProduct.booking_id == Booking.id)\
        .filter(
            BookingProduct.product_id == product.id,
            Booking.status == 'completed'
        )
    if supervisor_neighborhood_ids is not None:
        sold_quantity_query = sold_quantity_query.filter(
            Booking.neighborhood_id.in_(supervisor_neighborhood_ids)
        ) if supervisor_neighborhood_ids else sold_quantity_query.filter(Booking.id == -1)
    sold_quantity = sold_quantity_query.scalar() or 0
        
    # Calculate total revenue using unit_price if available, else product.price
    total_revenue_query = db.session.query(
        func.sum(
            BookingProduct.quantity * func.coalesce(BookingProduct.unit_price, Product.price)
        )
    ).join(Booking, BookingProduct.booking_id == Booking.id)\
     .join(Product, BookingProduct.product_id == Product.id)\
     .filter(
        BookingProduct.product_id == product.id,
        Booking.status == 'completed'
    )
    if supervisor_neighborhood_ids is not None:
        total_revenue_query = total_revenue_query.filter(
            Booking.neighborhood_id.in_(supervisor_neighborhood_ids)
        ) if supervisor_neighborhood_ids else total_revenue_query.filter(Booking.id == -1)
    total_revenue = total_revenue_query.scalar() or 0

    current_stock_query = db.session.query(func.sum(ProductStock.quantity)).filter(
        ProductStock.product_id == product.id,
        ProductStock.warehouse_id.isnot(None)
    )
    if current_user.role == 'supervisor':
        allowed_warehouse_ids = [warehouse.id for warehouse in _scoped_warehouses()]
        current_stock_query = current_stock_query.filter(
            ProductStock.warehouse_id.in_(allowed_warehouse_ids)
        ) if allowed_warehouse_ids else current_stock_query.filter(ProductStock.id == -1)
    current_stock = current_stock_query.scalar()
    if current_stock is None:
        current_stock = 0 if current_user.role == 'supervisor' else product.stock_quantity
    
    # Get recent bookings for this product
    recent_bookings_query = db.session.query(Booking, BookingProduct.quantity, BookingProduct.unit_price)\
        .join(BookingProduct, Booking.id == BookingProduct.booking_id)\
        .filter(BookingProduct.product_id == product.id)\
        .order_by(Booking.date.desc(), Booking.time.desc())
    if supervisor_neighborhood_ids is not None:
        recent_bookings_query = recent_bookings_query.filter(
            Booking.neighborhood_id.in_(supervisor_neighborhood_ids)
        ) if supervisor_neighborhood_ids else recent_bookings_query.filter(Booking.id == -1)
    recent_bookings = recent_bookings_query.limit(20).all()
        
    return render_template('admin/product_stats.html', product=product, sold_quantity=sold_quantity, total_revenue=total_revenue, current_stock=current_stock, recent_bookings=recent_bookings)




@bp.route('/locations/neighborhood/add/<int:city_id>', methods=['GET', 'POST'])
def add_neighborhood(city_id):
    city = City.query.get_or_404(city_id)
    
    if request.method == 'POST':
        name_ar = request.form.get('name_ar', '').strip()
        name_en = request.form.get('name_en', '').strip()
        osm_name = request.form.get('osm_name', '').strip()
        boundary_coords = request.form.get('boundary_coords', '').strip()
        
        if not name_ar or not name_en:
            flash('يجب إدخال اسم الحي بالعربية والإنجليزية', 'error')
            return redirect(url_for('admin.locations'))
        
        neighborhood = Neighborhood(
            city_id=city.id,
            name_ar=name_ar,
            name_en=name_en,
            osm_name=osm_name if osm_name else None,
            boundary_coords=boundary_coords if boundary_coords else None,
            latitude=request.form.get('latitude', type=float),
            longitude=request.form.get('longitude', type=float),
            is_active=True
        )
        db.session.add(neighborhood)
        db.session.commit()
        flash('تم إضافة الحي بنجاح', 'success')
        return redirect(url_for('admin.locations'))
    
    form = NeighborhoodForm()
    form.city_id.choices = [(c.id, c.name_ar) for c in City.query.all()]
    form.city_id.data = city.id
    return render_template('admin/location_form.html', form=form, title='إضافة حي', type='neighborhood')

@bp.route('/locations/neighborhood/edit/<int:id>', methods=['GET', 'POST'])
def edit_neighborhood(id):
    neighborhood = Neighborhood.query.get_or_404(id)
    
    if request.method == 'POST':
        neighborhood.name_ar = request.form.get('name_ar', neighborhood.name_ar).strip()
        neighborhood.name_en = request.form.get('name_en', neighborhood.name_en).strip()
        osm_name = request.form.get('osm_name', '').strip()
        neighborhood.osm_name = osm_name if osm_name else None
        boundary_coords = request.form.get('boundary_coords', '').strip()
        neighborhood.boundary_coords = boundary_coords if boundary_coords else None
        neighborhood.latitude = request.form.get('latitude', type=float)
        neighborhood.longitude = request.form.get('longitude', type=float)
        neighborhood.is_active = 'is_active' in request.form or request.form.get('is_active') == 'true'
        db.session.commit()
        flash('تم تعديل الحي', 'success')
        return redirect(url_for('admin.locations'))
    
    form = NeighborhoodForm(obj=neighborhood)
    form.city_id.choices = [(c.id, c.name_ar) for c in City.query.all()]
    return render_template('admin/location_form.html', form=form, title='تعديل حي', type='neighborhood', neighborhood=neighborhood)

@bp.route('/locations/city/delete/<int:id>', methods=['POST'])
def delete_city(id):
    city = City.query.get_or_404(id)
    # Check if city has neighborhoods
    if city.neighborhoods.count() > 0:
        flash('لا يمكن حذف المدينة لأنها تحتوي على أحياء. احذف الأحياء أولاً.')
        return redirect(url_for('admin.locations'))
    
    db.session.delete(city)
    db.session.commit()
    flash('تم حذف المدينة بنجاح')
    return redirect(url_for('admin.locations'))

@bp.route('/locations/neighborhood/delete/<int:id>', methods=['POST'])
def delete_neighborhood(id):
    neighborhood = Neighborhood.query.get_or_404(id)
    db.session.delete(neighborhood)
    db.session.commit()
    flash('تم حذف الحي بنجاح')
    return redirect(url_for('admin.locations'))

@bp.route('/api/neighborhood/<int:id>/boundary')
def get_neighborhood_boundary(id):
    neighborhood = Neighborhood.query.get_or_404(id)
    if neighborhood.boundary_coords:
        import json
        try:
            boundary = json.loads(neighborhood.boundary_coords)
            return jsonify({'boundary': boundary})
        except json.JSONDecodeError:
            pass
    return jsonify({'boundary': None})

# --- Subscription Package Management ---
@bp.route('/packages')
def packages():
    from sqlalchemy import func
    from app.models import Subscription
    
    packages = SubscriptionPackage.query.all()
    cities = City.query.filter_by(is_active=True).all()
    
    packages_data = []
    total_subscriptions = 0
    total_sales_revenue = 0
    
    for package in packages:
        subs_count = Subscription.query.filter_by(package_id=package.id).count()
        polishing_count = PolishingOrder.query.filter_by(package_id=package.id).count()
        # Estimate revenue since actual paid amount not stored
        rev = subs_count * (package.price or 0.0)
        
        packages_data.append({
            'package': package,
            'subs_count': subs_count,
            'polishing_count': polishing_count,
            'revenue': rev
        })
        total_subscriptions += subs_count
        total_sales_revenue += rev
        
    cities_json = [{'id': c.id, 'name_ar': c.name_ar} for c in cities]
    
    return render_template('admin/packages.html', 
                         packages=packages_data,
                         cities=cities,
                         cities_json=cities_json,
                         total_subscriptions=total_subscriptions,
                         total_sales_revenue=total_sales_revenue)

# --- City-Based Pricing: Packages ---
@bp.route('/api/city-package-prices/<int:package_id>')
def get_city_package_prices(package_id):
    """Get all city prices for a subscription package"""
    from app.models import CityPackagePrice
    prices = CityPackagePrice.query.filter_by(package_id=package_id).all()
    return jsonify([{
        'id': p.id,
        'city_id': p.city_id,
        'city_name': City.query.get(p.city_id).name_ar if City.query.get(p.city_id) else '',
        'price': p.price,
        'is_active': p.is_active
    } for p in prices])

@bp.route('/packages/assign-city', methods=['POST'])
def assign_package_to_city():
    """Assign a package to a city with a specific price"""
    from app.models import CityPackagePrice
    package_id = request.form.get('package_id', type=int)
    city_id = request.form.get('city_id', type=int)
    price = request.form.get('price', type=float)
    
    if not all([package_id, city_id, price is not None]):
        flash('بيانات غير مكتملة', 'error')
        return redirect(url_for('admin.packages'))
    
    existing = CityPackagePrice.query.filter_by(city_id=city_id, package_id=package_id).first()
    if existing:
        flash('الباقة مسندة لهذه المدينة بالفعل', 'error')
        return redirect(url_for('admin.packages'))
    
    cpp = CityPackagePrice(city_id=city_id, package_id=package_id, price=price, is_active=True)
    db.session.add(cpp)
    db.session.commit()
    flash('تم إسناد الباقة للمدينة بنجاح', 'success')
    return redirect(url_for('admin.packages'))

@bp.route('/packages/update-city-price', methods=['POST'])
def update_package_city_price():
    """Update city price for a package"""
    from app.models import CityPackagePrice
    price_id = request.form.get('price_id', type=int)
    new_price = request.form.get('price', type=float)
    
    cpp = CityPackagePrice.query.get_or_404(price_id)
    cpp.price = new_price
    db.session.commit()
    flash('تم تحديث السعر بنجاح', 'success')
    return redirect(url_for('admin.packages'))

@bp.route('/packages/remove-city-price/<int:price_id>', methods=['POST'])
def remove_package_city_price(price_id):
    """Remove a city price assignment for a package"""
    from app.models import CityPackagePrice
    cpp = CityPackagePrice.query.get_or_404(price_id)
    db.session.delete(cpp)
    db.session.commit()
    flash('تم إزالة الباقة من المدينة', 'success')
    return redirect(url_for('admin.packages'))

@bp.route('/packages/add', methods=['GET', 'POST'])
def add_package():
    form = SubscriptionPackageForm()
    if form.validate_on_submit():
        package = SubscriptionPackage(
            name_ar=form.name_ar.data,
            name_en=form.name_en.data,
            package_type=form.package_type.data,
            price=float(form.price.data),
            wash_count=int(form.wash_count.data),
            duration_days=int(form.duration_days.data),
            description=form.description.data,
            is_active=form.is_active.data
        )
        db.session.add(package)
        db.session.commit()
        flash('تم إضافة الباقة بنجاح')
        return redirect(url_for('admin.packages'))
    return render_template('admin/package_form.html', form=form, title='إضافة باقة')

@bp.route('/packages/edit/<int:id>', methods=['GET', 'POST'])
def edit_package(id):
    package = SubscriptionPackage.query.get_or_404(id)
    form = SubscriptionPackageForm()
    if form.validate_on_submit():
        package.name_ar = form.name_ar.data
        package.name_en = form.name_en.data
        package.package_type = form.package_type.data
        package.price = float(form.price.data)
        package.wash_count = int(form.wash_count.data)
        package.duration_days = int(form.duration_days.data)
        package.description = form.description.data
        package.is_active = form.is_active.data
        db.session.commit()
        flash('تم تعديل الباقة')
        return redirect(url_for('admin.packages'))
    elif request.method == 'GET':
        form.name_ar.data = package.name_ar
        form.name_en.data = package.name_en
        form.package_type.data = package.package_type or 'subscription'
        form.price.data = str(package.price)
        form.wash_count.data = str(package.wash_count)
        form.duration_days.data = str(package.duration_days)
        form.description.data = package.description
        form.is_active.data = package.is_active
    return render_template('admin/package_form.html', form=form, title='تعديل باقة')

@bp.route('/packages/delete/<int:id>', methods=['POST'])
def delete_package(id):
    package = SubscriptionPackage.query.get_or_404(id)
    db.session.delete(package)
    db.session.commit()
    flash('تم حذف الباقة')
    return redirect(url_for('admin.packages'))

# --- Subscription Requests Management ---
@bp.route('/subscriptions')
@login_required
def subscriptions():
    import json
    status = request.args.get('status', 'active')
    search_query = request.args.get('search', '').strip()
    
    subscriptions_query = Subscription.query
    
    # Filter by status
    if status != 'all':
        subscriptions_query = subscriptions_query.filter_by(status=status)
        
    supervisor_neighborhood_ids = []
    supervisor_city_ids = set()
    # Filter for supervisors
    if current_user.role == 'supervisor':
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
            supervisor_city_ids.update(n.city_id for n in current_user.supervisor_neighborhoods)
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_city_ids.add(city.id)
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        if supervisor_neighborhood_ids:
            subscriptions_query = subscriptions_query.filter(Subscription.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            subscriptions_query = subscriptions_query.filter_by(id=-1) # Empty result
        
    # Search filter
    if search_query:
        subscriptions_query = subscriptions_query.join(User, User.id == Subscription.customer_id).filter(
            (User.username.contains(search_query)) | 
            (User.phone.contains(search_query))
        )
    
    page = request.args.get('page', 1, type=int)
    pagination = subscriptions_query.order_by(Subscription.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    subscriptions_result = pagination.items
    
    # Get counts for tabs
    counts_query = Subscription.query
    if current_user.role == 'supervisor':
        counts_query = counts_query.filter(Subscription.neighborhood_id.in_(supervisor_neighborhood_ids)) if supervisor_neighborhood_ids else counts_query.filter(Subscription.id == -1)
    pending_count = counts_query.filter_by(status='pending').count()
    active_count = counts_query.filter_by(status='active').count()
    rejected_count = counts_query.filter_by(status='rejected').count()
    expired_count = counts_query.filter_by(status='expired').count()
    
    # Prepare JSON data for JavaScript
    subs_json = json.dumps([{
        'id': s.id,
        'customer_id': s.customer_id,
        'employee_id': s.employee_id,
        'neighborhood_id': s.neighborhood_id,
        'city_id': s.neighborhood.city_id if s.neighborhood else None,
        'remaining_washes': s.remaining_washes or 0,
        'end_date': s.end_date.isoformat() if s.end_date else None
    } for s in subscriptions_result])
    
    if current_user.role == 'supervisor':
        cities = City.query.filter(City.id.in_(supervisor_city_ids)).all() if supervisor_city_ids else []
    else:
        cities = City.query.all()
    cities_json = json.dumps([{
        'id': c.id,
        'name_ar': c.name_ar,
        'neighborhoods': [{'id': n.id, 'name_ar': n.name_ar} for n in c.neighborhoods if current_user.role != 'supervisor' or n.id in supervisor_neighborhood_ids]
    } for c in cities])
    
    employees = [emp for emp in _scoped_employee_query(include_break=True).all() if not _is_employee_on_break(emp)]
    if current_user.role == 'supervisor':
        customers = User.query.filter_by(role='customer').join(Booking, User.id == Booking.customer_id).filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids)).distinct().all() if supervisor_neighborhood_ids else []
    else:
        customers = User.query.filter_by(role='customer').all()
    packages = SubscriptionPackage.query.filter_by(is_active=True, package_type='subscription').all()
    
    return render_template('admin/subscriptions.html',
                          subscriptions=subscriptions_result,
                          current_status=status,
                          search_query=search_query,
                          pending_count=pending_count,
                          active_count=active_count,
                          rejected_count=rejected_count,
                          expired_count=expired_count,
                          subscriptions_json=subs_json,
                          cities=cities,
                          cities_json=cities_json,
                          employees=employees,
                          customers=customers,
                          packages=packages,
                          pagination=pagination)

@bp.route('/subscriptions/create', methods=['POST'])
def create_subscription():
    from datetime import datetime, timedelta
    
    customer_id = request.form.get('customer_id')
    package_id = request.form.get('package_id')
    employee_id = request.form.get('employee_id')
    neighborhood_id = request.form.get('neighborhood_id')
    discount = float(request.form.get('discount', 0))

    # Validate supervisor scope
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        # Check if the neighborhood is within supervisor's scope
        if int(neighborhood_id) not in supervisor_neighborhood_ids:
            flash('خطأ: لا يمكنك إضافة اشتراك خارج نطاق منطقتك المحددة', 'error')
            return redirect(url_for('admin.subscriptions'))
    
    package = SubscriptionPackage.query.get(package_id)
    if not package or package.package_type != 'subscription':
        flash('الباقة غير موجودة')
        return redirect(url_for('admin.subscriptions'))

    if employee_id:
        employee = User.query.get(int(employee_id))
        if not employee or employee.role != 'employee' or _is_employee_on_break(employee):
            flash('لا يمكن إسناد الاشتراك لموظف في وضع الراحة', 'error')
            return redirect(url_for('admin.subscriptions'))
    
    # Create subscription (employee is now optional - booking will find available employee)
    subscription = Subscription(
        customer_id=int(customer_id),
        employee_id=int(employee_id) if employee_id else None,
        neighborhood_id=int(neighborhood_id),
        package_id=int(package_id),
        remaining_washes=int(package.wash_count),
        start_date=datetime.now().date(),
        end_date=(datetime.now() + timedelta(days=int(package.duration_days))).date(),
        status='active'
    )
    
    db.session.add(subscription)
    db.session.commit()
    
    flash(f'تم إضافة الاشتراك بنجاح (الخصم: {discount}%)')
    return redirect(url_for('admin.subscriptions', status='active'))

@bp.route('/subscriptions/<int:id>/approve', methods=['POST'])
def approve_subscription(id):
    subscription = Subscription.query.get_or_404(id)
    employee_id = request.form.get('employee_id')
    
    # Employee is now optional - bookings will find available employee in neighborhood
    subscription.status = 'active'
    if employee_id:
        employee = User.query.get(int(employee_id))
        if not employee or employee.role != 'employee' or employee.is_on_break:
            flash('لا يمكن إسناد الاشتراك لموظف في وضع الراحة', 'error')
            return redirect(url_for('admin.subscriptions'))
        subscription.employee_id = int(employee_id)
    
    # Set remaining washes from package
    if subscription.package:
        subscription.remaining_washes = int(subscription.package.wash_count)
    
    db.session.commit()
    flash('تم قبول الاشتراك بنجاح')
    return redirect(url_for('admin.subscriptions', status='active'))

@bp.route('/subscriptions/<int:id>/reject', methods=['POST'])
def reject_subscription(id):
    subscription = Subscription.query.get_or_404(id)
    subscription.status = 'rejected'
    db.session.commit()
    flash('تم رفض الطلب')
    return redirect(url_for('admin.subscriptions', status='rejected'))

@bp.route('/subscriptions/<int:id>/reassign', methods=['POST'])
def reassign_subscription(id):
    subscription = Subscription.query.get_or_404(id)
    employee_id = request.form.get('employee_id')
    
    if not employee_id:
        flash('يجب اختيار موظف')
        return redirect(url_for('admin.subscriptions', status='active'))

    employee = User.query.get(int(employee_id))
    if not employee or employee.role != 'employee' or employee.is_on_break:
        flash('لا يمكن إسناد الاشتراك لموظف في وضع الراحة', 'error')
        return redirect(url_for('admin.subscriptions', status='active'))
    
    subscription.employee_id = int(employee_id)
    db.session.commit()
    flash('تم إعادة إسناد الاشتراك بنجاح')
    return redirect(url_for('admin.subscriptions', status='active'))

@bp.route('/subscriptions/<int:id>/edit', methods=['POST'])
def edit_subscription(id):
    subscription = Subscription.query.get_or_404(id)
    
    # Update employee
    employee_id = request.form.get('employee_id')
    if employee_id:
        employee = User.query.get(int(employee_id))
        if not employee or employee.role != 'employee' or employee.is_on_break:
            flash('لا يمكن إسناد الاشتراك لموظف في وضع الراحة', 'error')
            return redirect(url_for('admin.subscriptions', status='active'))
        subscription.employee_id = int(employee_id)
    
    # Update location
    neighborhood_id = request.form.get('neighborhood_id')
    if neighborhood_id:
        subscription.neighborhood_id = int(neighborhood_id)
    
    # Update remaining washes
    remaining_washes = request.form.get('remaining_washes')
    if remaining_washes:
        subscription.remaining_washes = int(remaining_washes)
    
    # Update end date
    end_date_str = request.form.get('end_date')
    if end_date_str:
        from datetime import datetime
        subscription.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    db.session.commit()
    flash('تم تحديث الاشتراك بنجاح')
    return redirect(url_for('admin.subscriptions', status='active'))

@bp.route('/subscriptions/<int:id>/delete', methods=['POST'])
def delete_subscription(id):
    subscription = Subscription.query.get_or_404(id)
    db.session.delete(subscription)
    db.session.commit()
    flash('تم حذف الاشتراك بنجاح')
    return redirect(url_for('admin.subscriptions', status='active'))

@bp.route('/subscriptions/whatsapp/<int:id>')
def whatsapp_customer(id):
    subscription = Subscription.query.get_or_404(id)
    customer = subscription.customer
    
    # Format phone number for WhatsApp (remove + and spaces)
    phone = customer.phone.replace('+', '').replace(' ', '') if customer.phone else ''
    
    # Create message
    message = f"مرحباً {customer.username}، نود التواصل معك بخصوص طلب الاشتراك رقم #{subscription.id}"
    
    # WhatsApp URL
    whatsapp_url = f"https://wa.me/{phone}?text={quote(message)}"
    
    return redirect(whatsapp_url)

# --- Polishing Requests Management ---
@bp.route('/polishing-orders')
@login_required
def polishing_orders():
    status = request.args.get('status', 'pending')
    search_query = request.args.get('search', '').strip()

    orders_query = PolishingOrder.query

    if status != 'all':
        orders_query = orders_query.filter_by(status=status)

    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])

        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])

        if supervisor_neighborhood_ids:
            orders_query = orders_query.filter(PolishingOrder.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            orders_query = orders_query.filter_by(id=-1)

    if search_query:
        orders_query = orders_query.join(User, User.id == PolishingOrder.customer_id).filter(
            (User.username.contains(search_query)) |
            (User.phone.contains(search_query))
        )

    page = request.args.get('page', 1, type=int)
    pagination = orders_query.order_by(PolishingOrder.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    orders = pagination.items

    pending_count = PolishingOrder.query.filter_by(status='pending').count()
    accepted_count = PolishingOrder.query.filter_by(status='accepted').count()
    completed_count = PolishingOrder.query.filter_by(status='completed').count()
    rejected_count = PolishingOrder.query.filter_by(status='rejected').count()

    return render_template(
        'admin/polishing_orders.html',
        orders=orders,
        current_status=status,
        search_query=search_query,
        pending_count=pending_count,
        accepted_count=accepted_count,
        completed_count=completed_count,
        rejected_count=rejected_count,
        pagination=pagination
    )

@bp.route('/polishing-orders/<int:id>/accept', methods=['POST'])
def accept_polishing_order(id):
    order = PolishingOrder.query.get_or_404(id)
    order.status = 'accepted'
    db.session.commit()
    flash('تم قبول طلب التلميع', 'success')
    return redirect(request.referrer or url_for('admin.polishing_orders', status='accepted'))

@bp.route('/polishing-orders/<int:id>/complete', methods=['POST'])
def complete_polishing_order(id):
    order = PolishingOrder.query.get_or_404(id)
    order.status = 'completed'
    db.session.commit()
    flash('تم إكمال طلب التلميع', 'success')
    return redirect(request.referrer or url_for('admin.polishing_orders', status='completed'))

@bp.route('/polishing-orders/<int:id>/reject', methods=['POST'])
def reject_polishing_order(id):
    order = PolishingOrder.query.get_or_404(id)
    order.status = 'rejected'
    db.session.commit()
    flash('تم رفض طلب التلميع', 'success')
    return redirect(request.referrer or url_for('admin.polishing_orders', status='rejected'))

@bp.route('/polishing-orders/<int:id>/delete', methods=['POST'])
def delete_polishing_order(id):
    order = PolishingOrder.query.get_or_404(id)
    db.session.delete(order)
    db.session.commit()
    flash('تم حذف طلب التلميع', 'success')
    return redirect(request.referrer or url_for('admin.polishing_orders'))

# API endpoint for getting employees by neighborhood
@bp.route('/api/employees-by-neighborhood/<int:neighborhood_id>')
def employees_by_neighborhood(neighborhood_id):
    from app.models import User, employee_neighborhoods
    
    # Get employees assigned to this neighborhood
    employees = User.query.join(employee_neighborhoods).filter(
        employee_neighborhoods.c.neighborhood_id == neighborhood_id,
        User.role == 'employee'
    ).all()
    employees = [emp for emp in employees if not _is_employee_on_break(emp)]
    
    return jsonify([{'id': emp.id, 'username': emp.username} for emp in employees])

# --- Booking Management ---
BOOKING_NEXT_STATUS = {
    'assigned': 'en_route',
    'en_route': 'arrived',
    'arrived': 'in_progress',
    'in_progress': 'completed',
}

BOOKING_STATUS_LABELS = {
    'assigned': 'تم التعيين',
    'en_route': 'في الطريق',
    'arrived': 'وصل الموظف',
    'in_progress': 'جاري العمل',
    'completed': 'مكتمل',
}


def _booking_service_display_name(booking):
    if booking.subscription_id and booking.subscription and booking.subscription.package:
        return booking.subscription.package.name_ar

    first_item = booking.items.first()
    if first_item and first_item.service:
        return first_item.service.name_ar

    if booking.service:
        return booking.service.name_ar

    return 'الخدمة'


def _booking_awards_loyalty_point(booking):
    items = booking.items.all()
    if items:
        return any(
            item.service and item.service.awards_loyalty_point is not False
            for item in items
        )

    return bool(booking.service and booking.service.awards_loyalty_point is not False)


def _apply_booking_completion_effects(booking):
    from app.models import ReferralRecord

    if not booking.customer:
        return 'تم إكمال الخدمة بنجاح'

    message = 'تم إكمال الخدمة بنجاح'

    if not booking.subscription_id and not booking.used_free_wash:
        if _booking_awards_loyalty_point(booking):
            if booking.customer.add_loyalty_point():
                message = 'تم إكمال الخدمة. وصل العميل للحد المطلوب وحصل على غسلة مجانية!'
            else:
                message = 'تم إكمال الخدمة وإضافة نقطة ولاء للعميل'
        else:
            message = 'تم إكمال الخدمة بدون إضافة نقطة ولاء'

    first_completed_count = Booking.query.filter_by(
        customer_id=booking.customer_id,
        status='completed'
    ).count()

    if first_completed_count <= 1:
        if booking.customer.used_influencer_code_id:
            if booking.customer.add_loyalty_point():
                db.session.add(Notification(
                    user_id=booking.customer_id,
                    title='حصلت على غسلة مجانية!',
                    message='لقد حصلت على غسلة مجانية جديدة!',
                    created_at=datetime.utcnow()
                ))
            else:
                db.session.add(Notification(
                    user_id=booking.customer_id,
                    title='نقطة إضافية',
                    message='حصلت على نقطة إضافية لتسجيلك باستخدام كود مؤثر!',
                    created_at=datetime.utcnow()
                ))

        referral_record = ReferralRecord.query.filter_by(
            referred_user_id=booking.customer_id,
            first_wash_completed=False
        ).first()

        if referral_record:
            referral_record.first_wash_completed = True
            referral_record.completed_at = datetime.utcnow()

            site_settings = SiteSettings.get_settings()
            target = site_settings.referral_target_count or 10
            completed_referrals = ReferralRecord.query.filter_by(
                referrer_id=referral_record.referrer_id,
                first_wash_completed=True
            ).count()

            if completed_referrals > 0 and completed_referrals % target == 0:
                referrer = User.query.get(referral_record.referrer_id)
                if referrer:
                    referrer.free_washes = (referrer.free_washes or 0) + 1
                    db.session.add(Notification(
                        user_id=referrer.id,
                        title='حصلت على غسلة مجانية!',
                        message=f'مبروك! أكمل {target} من أصدقائك المحالين غسلتهم الأولى. تمت إضافة غسلة مجانية لحسابك!',
                        created_at=datetime.utcnow()
                    ))

    for booking_product in booking.products:
        product = booking_product.product
        if product and product.stock_quantity is not None:
            product.stock_quantity -= booking_product.quantity
            if product.stock_quantity < 0:
                product.stock_quantity = 0

    return message


def _send_booking_status_notification(booking, status):
    if not booking.customer:
        return

    from app.notifications import send_push_notification

    service_name = _booking_service_display_name(booking)
    status_messages = {
        'en_route': {
            'title': 'الموظف في الطريق',
            'body': f'موظفنا في الطريق إليك! سيصل قريباً لحجزك #{booking.id}'
        },
        'arrived': {
            'title': 'وصل الموظف',
            'body': f'وصل موظفنا إلى موقعك للحجز #{booking.id}'
        },
        'in_progress': {
            'title': 'جاري العمل',
            'body': f'بدأ موظفنا بتقديم خدمة {service_name} للحجز #{booking.id}'
        }
    }

    if status not in status_messages:
        return

    send_push_notification(booking.customer, {
        "title": status_messages[status]['title'],
        "body": status_messages[status]['body'],
        "icon": "/static/images/logo.png",
        "badge": "/static/images/logo.png",
        "url": "/customer/bookings",
        "data": {
            "booking_id": booking.id,
            "status": status
        }
    })


def _send_booking_rating_request(booking):
    if not booking.customer:
        return

    try:
        from app.notifications import send_push_notification

        notification = Notification(
            user_id=booking.customer_id,
            title='تم الانتهاء من الغسيل!',
            message='نأمل أن تكون راضياً عن خدمتنا. يرجى تقييم تجربتك.',
            created_at=datetime.utcnow()
        )
        db.session.add(notification)
        db.session.commit()

        send_push_notification(
            booking.customer,
            {
                "title": 'تم الانتهاء من الغسيل!',
                "body": 'نأمل أن تكون راضياً عن خدمتنا. يرجى تقييم تجربتك.',
                "url": url_for('customer.rate_booking', booking_id=booking.id, _external=True)
            }
        )
    except Exception as e:
        print(f"Error sending rating notification: {e}")


@bp.route('/bookings')
def bookings():
    import json
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    employee_filter = request.args.get('employee', 'all')
    from_date_filter = request.args.get('from_date', '')
    to_date_filter = request.args.get('to_date', '')
    sort_dir = request.args.get('sort', 'desc')  # asc or desc
    search_query = request.args.get('q', '').strip()
    
    query = Booking.query
    
    # Filter for supervisors
    supervisor_neighborhood_ids = []
    if current_user.role == 'supervisor':
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        if supervisor_neighborhood_ids:
            query = query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            query = query.filter_by(id=-1)  # Empty result
    
    if status_filter == 'current':
        query = query.filter(~Booking.status.in_(['completed', 'cancelled']))
    elif status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if employee_filter != 'all':
        query = query.filter_by(employee_id=int(employee_filter))
    
    # Date range filter
    from datetime import datetime
    if from_date_filter:
        try:
            from_date = datetime.strptime(from_date_filter, '%Y-%m-%d').date()
            query = query.filter(Booking.date >= from_date)
        except ValueError:
            pass
    if to_date_filter:
        try:
            to_date = datetime.strptime(to_date_filter, '%Y-%m-%d').date()
            query = query.filter(Booking.date <= to_date)
        except ValueError:
            pass
        
    # Apply search
    if search_query:
        # Check if search query is a number (for ID search)
        if search_query.isdigit():
             query = query.join(User, Booking.customer_id == User.id).filter(
                or_(
                    Booking.id == int(search_query),
                    User.phone.ilike(f'%{search_query}%')
                )
            )
        else:
            query = query.join(User, Booking.customer_id == User.id).filter(
                User.username.ilike(f'%{search_query}%')
            )
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Order by date and time
    if sort_dir == 'asc':
        pagination = query.order_by(Booking.date.asc(), Booking.time.asc()).paginate(page=page, per_page=per_page, error_out=False)
    else:
        pagination = query.order_by(Booking.date.desc(), Booking.time.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    bookings_list = pagination.items
    
    # Filter cities and neighborhoods for supervisor
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            # Get cities that contain the supervisor's neighborhoods
            city_ids = set()
            for neighborhood in Neighborhood.query.filter(Neighborhood.id.in_(supervisor_neighborhood_ids)).all():
                city_ids.add(neighborhood.city_id)
            
            cities = City.query.filter(City.id.in_(city_ids)).all()
        else:
            cities = []
    else:
        cities = City.query.all()
    
    cities_json = json.dumps([{
        'id': c.id,
        'name_ar': c.name_ar,
        'neighborhoods': [{'id': n.id, 'name_ar': n.name_ar} for n in c.neighborhoods if current_user.role != 'supervisor' or n.id in supervisor_neighborhood_ids]
    } for c in cities])
    
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Filter employees for supervisor
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            employees = User.query.filter_by(role='employee').join(User.neighborhoods).filter(
                Neighborhood.id.in_(supervisor_neighborhood_ids)
            ).distinct().all()
        else:
            employees = []
    else:
        employees = User.query.filter_by(role='employee').all()
    employees = [emp for emp in employees if not _is_employee_on_break(emp)]
    
    customers = User.query.filter_by(role='customer').all()
    services_query = Service.query.all()
    services = []
    for s in services_query:
        # Include city-size prices for frontend logic
        # Format: { city_id: { size_id: price } }
        city_size_prices_map = {}
        for cp in s.city_size_prices:
            c_id = str(cp.city_id)
            if c_id not in city_size_prices_map:
                city_size_prices_map[c_id] = {}
            city_size_prices_map[c_id][cp.vehicle_size_id] = float(cp.price)
            
        services.append({
            'id': s.id,
            'name_ar': s.name_ar,
            'price': float(s.price),
            'city_size_prices': city_size_prices_map
        })
    
    vehicle_sizes = VehicleSize.query.all()
    vehicle_sizes_json = json.dumps({vs.id: float(vs.price_adjustment or 0) for vs in vehicle_sizes})
    
    # Get counts for status tabs
    current_statuses = ['pending', 'assigned', 'en_route', 'arrived', 'in_progress']
    counts_query = Booking.query
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            counts_query = counts_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            counts_query = counts_query.filter(Booking.id == -1)
    current_count = counts_query.filter(Booking.status.in_(current_statuses)).count()
    completed_count = counts_query.filter_by(status='completed').count()
    cancelled_count = counts_query.filter_by(status='cancelled').count()
    
    return render_template('admin/bookings.html', bookings=bookings_list, employees=employees, 
                           status_filter=status_filter, employee_filter=employee_filter,
                           from_date=from_date_filter, to_date=to_date_filter, sort_dir=sort_dir,
                           customers=customers, services=services, cities=cities, cities_json=cities_json, 
                           vehicle_sizes_json=vehicle_sizes_json, today=today,
                           current_count=current_count, completed_count=completed_count, cancelled_count=cancelled_count,
                           pagination=pagination)

@bp.route('/bookings/create', methods=['POST'])
def create_booking():
    from datetime import datetime, time as dt_time
    
    customer_id = request.form.get('customer_id')
    service_id = request.form.get('service_id')
    employee_id = request.form.get('employee_id')
    neighborhood_id = request.form.get('neighborhood_id')
    date = request.form.get('date')
    time_str = request.form.get('time')
    discount = float(request.form.get('discount', 0))

    customer = User.query.filter_by(id=customer_id, role='customer').first()
    if not customer:
        flash('اختر عميلاً صحيحاً من نتائج البحث', 'error')
        return redirect(url_for('admin.bookings'))

    booking_date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    hour, minute = map(int, time_str.split(':'))
    time_obj = dt_time(hour, minute)

    if employee_id:
        employee = User.query.get(int(employee_id))
        if not employee or employee.role != 'employee' or _is_employee_on_break(employee, booking_date_obj, time_obj):
            flash('لا يمكن إسناد الحجز لموظف في وضع الراحة', 'error')
            return redirect(url_for('admin.bookings'))
    
    # Validate supervisor scope
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        # Check if the neighborhood is within supervisor's scope
        if int(neighborhood_id) not in supervisor_neighborhood_ids:
            flash('خطأ: لا يمكنك إضافة حجز خارج نطاق منطقتك المحددة', 'error')
            return redirect(url_for('admin.bookings'))
    
    # Calculate base service price (considering seasons)
    service = Service.query.get(int(service_id))
    base_price = service.price if service else 0.0
    
    from app.models import Season
    active_season = Season.query.filter(
        Season.is_active == True,
        Season.start_date <= booking_date_obj,
        Season.end_date >= booking_date_obj
    ).first()
    
    if active_season and service:
        ssp = active_season.service_prices.filter_by(service_id=service.id).first()
        if ssp:
            base_price = ssp.price
            
    # Calculate custom service price with discount if discount > 0
    # Always set custom_service_price if there's a seasonal price or discount to lock it
    
    # 1. City & Size Specific Price (Unified Pricing)
    vehicle_id = request.form.get('vehicle_id')
    vehicle_size_price = 0.0
    
    if vehicle_id and neighborhood_id:
        from app.models import Vehicle, CityServicePrice, Neighborhood
        vehicle = Vehicle.query.get(int(vehicle_id))
        if not vehicle or vehicle.user_id != customer.id:
            flash('السيارة المحددة لا تخص العميل المختار', 'error')
            return redirect(url_for('admin.bookings'))
        neighborhood_obj = Neighborhood.query.get(int(neighborhood_id))
        
        if vehicle and neighborhood_obj:
            city_id = neighborhood_obj.city_id
            # Query the unified table
            csp = CityServicePrice.query.filter_by(
                city_id=city_id,
                service_id=service.id,
                vehicle_size_id=vehicle.vehicle_size_id
            ).first()
            
            if csp:
                # Unified price overrides the base service price
                base_price = csp.price
                vehicle_size_price = 0.0 # Price is already unified
            elif vehicle.size:
                # Fallback to general size adjustment if no specific city-size override
                vehicle_size_price = vehicle.size.price_adjustment

    custom_service_price = base_price
    if discount > 0:
        discount_amount = base_price * (discount / 100.0)
        custom_service_price = base_price - discount_amount

    # If employee is assigned, status should be 'assigned', otherwise 'pending'
    booking_status = 'assigned' if employee_id else 'pending'
    
    booking = Booking(
        customer_id=int(customer_id), 
        service_id=int(service_id), 
        employee_id=int(employee_id) if employee_id else None,
        neighborhood_id=int(neighborhood_id) if neighborhood_id else None,
        date=booking_date_obj, 
        time=time_obj,
        status=booking_status,
        custom_service_price=custom_service_price,
        vehicle_size_price=vehicle_size_price, 
        total_price=custom_service_price + vehicle_size_price
    )
    db.session.add(booking)
    db.session.flush()

    # Create BookingItem for consistency
    from app.models import BookingItem
    if vehicle_id:
        b_item = BookingItem(
            booking_id=booking.id,
            vehicle_id=int(vehicle_id),
            service_id=int(service_id),
            quantity=1,
            service_price=custom_service_price,
            size_price_adjustment=vehicle_size_price,
            total_item_price=custom_service_price + vehicle_size_price
        )
        db.session.add(b_item)
        booking.vehicle_id = int(vehicle_id)

    db.session.commit()
    
    # Notify employee if assigned
    if employee_id:
        employee = User.query.get(int(employee_id))
        if employee:
            from app.notifications import send_push_notification
            notification_data = {
                "title": "حجز جديد تم تعيينه لك 🆕",
                "body": f"تم تعيين حجز جديد #{booking.id}\nالعميل: {booking.customer.username}\nالخدمة: {booking.service.name_ar}\nالموعد: {booking.date} {booking.time.strftime('%H:%M')}",
                "icon": "/static/images/logo.png",
                "badge": "/static/images/logo.png",
                "url": "/employee/bookings/active",
                "data": {
                    "booking_id": booking.id
                }
            }
            send_push_notification(employee, notification_data)
            
    flash(f'تم إضافة الحجز بنجاح (الخصم: {discount}%)')
    return redirect(url_for('admin.bookings'))

@bp.route('/bookings/<int:booking_id>/delete-item/<int:item_id>', methods=['POST'])
def delete_booking_item(booking_id, item_id):
    booking = Booking.query.get_or_404(booking_id)
    from app.models import BookingItem
    item = BookingItem.query.get_or_404(item_id)
    
    if item.booking_id != booking.id:
        flash('العنصر غير تابع لهذا الحجز')
        return redirect(url_for('admin.bookings'))
    
    # Check if it's the last item
    if booking.items.count() <= 1:
        flash('لا يمكن حذف آخر مركبة من الحجز. يمكنك إلغاء الحجز بالكامل بدلاً من ذلك.', 'warning')
        return redirect(url_for('admin.bookings'))
    
    db.session.delete(item)
    db.session.commit()
    
    # Recalculate total price
    update_booking_totals(booking)
    
    flash('تم حذف المركبة من الحجز وتعديل السعر الإجمالي')
    return redirect(url_for('admin.bookings'))

def update_booking_totals(booking):
    """Helper to recalculate total_price based on items and products"""
    total_services = sum((item.total_item_price or 0) for item in booking.items)
    total_products = sum(
        ((p.unit_price if p.unit_price is not None else (p.product.price if p.product else 0)) or 0) * (p.quantity or 0)
        for p in booking.products
    )
    if booking.items.count() == 0 and not booking.subscription_id and not booking.used_free_wash:
        service_price = booking.custom_service_price if booking.custom_service_price is not None else (booking.service.price if booking.service else 0)
        total_services = (service_price or 0) + (booking.vehicle_size_price or 0)
    
    # Apply discount if any (from header)
    if booking.discount_code:
        if booking.discount_code.discount_type == 'percentage':
            # Apply to services only?
            discount = (total_services * booking.discount_code.value) / 100
        else:
            discount = booking.discount_code.value
        booking.total_price = max(0, (total_services + total_products) - discount)
    else:
        booking.total_price = total_services + total_products
    
    db.session.commit()

@bp.route('/bookings/<int:id>/edit', methods=['POST'])
def edit_booking(id):
    booking = Booking.query.get_or_404(id)
    if not _booking_has_scope_access(booking):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062a\u0639\u062f\u064a\u0644 \u0647\u0630\u0627 \u0627\u0644\u062d\u062c\u0632', 'error')
        return redirect(request.referrer or url_for('admin.bookings'))

    new_date = request.form.get('date')
    new_time = request.form.get('time')
    current_date = booking.date.strftime('%Y-%m-%d')
    current_time = booking.time.strftime('%H:%M')
    appointment_changed = (
        (new_date is not None or new_time is not None)
        and (new_date != current_date or new_time != current_time)
    )
    if appointment_changed:
        if not booking.employee_id:
            flash('يجب إسناد الحجز إلى موظف قبل تغيير الموعد', 'error')
            return redirect(request.referrer or url_for('admin.bookings'))
        try:
            new_date_obj = datetime.strptime(new_date, '%Y-%m-%d').date()
            new_time_obj = datetime.strptime(new_time, '%H:%M').time()
        except (TypeError, ValueError):
            flash('التاريخ أو الوقت المحدد غير صحيح', 'error')
            return redirect(request.referrer or url_for('admin.bookings'))

        available_slots = _employee_available_slots(
            booking.employee_id,
            new_date_obj,
            booking.service_id,
            exclude_booking_id=booking.id
        )
        if new_time_obj.strftime('%H:%M') not in available_slots:
            flash('الوقت المحدد غير متاح. اختر وقتاً من قائمة الأوقات المتاحة', 'error')
            return redirect(request.referrer or url_for('admin.bookings'))

        booking.date = new_date_obj
        booking.time = new_time_obj

    has_item_fields = any(
        key.startswith('item_service_price_')
        for key in request.form.keys()
    )
    
    if has_item_fields:
        for item in booking.items.all():
            service_price = request.form.get(f'item_service_price_{item.id}')

            try:
                if service_price is not None and service_price.strip():
                    item.service_price = max(0.0, float(service_price))
            except ValueError:
                pass

            item.quantity = 1
            item.total_item_price = (item.service_price or 0.0) + (item.size_price_adjustment or 0.0)

        if booking.items.count() == 1:
            item = booking.items.first()
            booking.custom_service_price = item.service_price
            booking.vehicle_size_price = item.size_price_adjustment
        else:
            booking.custom_service_price = None
            booking.vehicle_size_price = 0.0
    else:
        # Legacy fallback for old bookings without BookingItem rows.
        try:
            vehicle_size_price = float(request.form.get('vehicle_size_price', booking.vehicle_size_price or 0))
            booking.vehicle_size_price = vehicle_size_price
        except ValueError:
            pass

        try:
            custom_service_price = request.form.get('custom_service_price')
            if custom_service_price and custom_service_price.strip():
                booking.custom_service_price = float(custom_service_price)
            elif 'custom_service_price' in request.form:
                booking.custom_service_price = None
        except ValueError:
            pass

        if booking.items.count() == 1:
            item = booking.items.first()
            item.service_price = booking.custom_service_price if booking.custom_service_price is not None else (booking.service.price if booking.service else 0.0)
            item.size_price_adjustment = booking.vehicle_size_price or 0.0
            item.quantity = 1
            item.total_item_price = item.service_price + item.size_price_adjustment
        
    # Update payment method
    payment_method = request.form.get('payment_method')
    if payment_method in ['cash', 'card']:
        booking.payment_method = payment_method

    # Update product prices
    from app.models import BookingProduct
    for key, value in request.form.items():
        if key.startswith('product_price_'):
            try:
                product_id = int(key.split('_')[2])
                price = float(value)
                
                # Find the booking product
                bp_item = BookingProduct.query.filter_by(booking_id=id, product_id=product_id).first()
                if bp_item:
                    bp_item.unit_price = price
            except (ValueError, IndexError):
                continue

        if key.startswith('product_quantity_'):
            try:
                product_id = int(key.split('_')[2])
                new_quantity = max(1, int(value))

                bp_item = BookingProduct.query.filter_by(booking_id=id, product_id=product_id).first()
                if bp_item:
                    old_quantity = bp_item.quantity or 0
                    diff = new_quantity - old_quantity

                    available_stock = (bp_item.product.stock_quantity or 0) if bp_item.product else 0
                    if diff > 0 and bp_item.product and available_stock < diff:
                        flash(f'الكمية المطلوبة غير متوفرة للمنتج "{bp_item.product.name_ar}". المتوفر: {available_stock}', 'error')
                        return redirect(request.referrer or url_for('admin.bookings'))

                    bp_item.quantity = new_quantity
                    if bp_item.product:
                        bp_item.product.stock_quantity = (bp_item.product.stock_quantity or 0) - diff
            except (ValueError, IndexError):
                continue
        
    update_booking_totals(booking)
    
    flash('تم تحديث الحجز بنجاح', 'success')
    return redirect(request.referrer or url_for('admin.bookings'))

@bp.route('/bookings/<int:id>/refund-product/<int:product_id>', methods=['POST'])
def refund_product(id, product_id):
    from app.models import BookingProduct, Product
    
    booking = Booking.query.get_or_404(id)
    if not _booking_has_scope_access(booking):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062a\u0639\u062f\u064a\u0644 \u0647\u0630\u0627 \u0627\u0644\u062d\u062c\u0632', 'error')
        return redirect(request.referrer or url_for('admin.bookings'))

    product_link = BookingProduct.query.filter_by(booking_id=id, product_id=product_id).first()
    
    if product_link:
        # Return to stock
        product = Product.query.get(product_id)
        if product:
            product.stock_quantity += product_link.quantity
            
        # Delete link
        db.session.delete(product_link)
    db.session.commit()
    
    # Recalculate total price
    update_booking_totals(booking)
    
    flash(f'تم استرجاع المنتج "{product.name_ar}" للمخزون بنجاح', 'success')
    return redirect(request.referrer or url_for('admin.bookings'))

@bp.route('/bookings/<int:id>/add-product', methods=['POST'])
def add_booking_product(id):
    from app.models import BookingProduct, Product
    
    booking = Booking.query.get_or_404(id)
    if not _booking_has_scope_access(booking):
        flash('\u0644\u0627 \u062a\u0645\u0644\u0643 \u0635\u0644\u0627\u062d\u064a\u0629 \u062a\u0639\u062f\u064a\u0644 \u0647\u0630\u0627 \u0627\u0644\u062d\u062c\u0632', 'error')
        return redirect(request.referrer or url_for('admin.bookings'))

    product_id = request.form.get('product_id')
    quantity = int(request.form.get('quantity', 1))
    
    if not product_id:
        flash('الرجاء اختيار منتج', 'error')
        return redirect(url_for('admin.bookings'))
        
    product = Product.query.get_or_404(product_id)
    
    # Check stock
    if product.stock_quantity < quantity:
        flash(f'الكمية المطلوبة غير متوفرة. المتوفر: {product.stock_quantity}', 'error')
        return redirect(url_for('admin.bookings'))
        
    # Check if product already in booking
    existing = BookingProduct.query.filter_by(booking_id=id, product_id=product_id).first()
    
    if existing:
        existing.quantity += quantity
    else:
        new_item = BookingProduct(
            booking_id=id,
            product_id=product_id,
            quantity=quantity,
            unit_price=product.price # Lock price at time of addition
        )
        db.session.add(new_item)
        
    # Deduct from stock
    product.stock_quantity -= quantity
    
    db.session.commit()
    
    # Recalculate total price
    update_booking_totals(booking)
    
    flash('تم إضافة المنتج للحجز بنجاح', 'success')
    
    # Redirect back to referring page (could be bookings or employee stats)
    return redirect(request.referrer or url_for('admin.bookings'))

@bp.route('/api/products/available')
def get_available_products_api():
    products = Product.query.filter(Product.stock_quantity > 0).all()
    return jsonify([{
        'id': p.id,
        'name': p.name_ar,
        'price': p.price,
        'stock': p.stock_quantity
    } for p in products])

# --- APIs ---

@bp.route('/bookings/<int:id>/items')
def get_booking_items_api(id):
    booking = Booking.query.get_or_404(id)
    if not _booking_has_scope_access(booking):
        return jsonify([]), 403

    items = []

    for item in booking.items.all():
        service_price = item.service_price if item.service_price is not None else (item.service.price if item.service else 0)

        if (booking.subscription_id or booking.used_free_wash) and (item.total_item_price or 0) == 0:
            service_price = 0

        items.append({
            'item_id': item.id,
            'vehicle': f"{item.vehicle.brand} ({item.vehicle.plate_number})" if item.vehicle else 'N/A',
            'service': item.service.name_ar if item.service else 'N/A',
            'service_price': service_price,
            'total': item.total_item_price or 0
        })

    return jsonify(items)

@bp.route('/bookings/<int:id>/products')
def get_booking_products_api(id):
    booking = Booking.query.get_or_404(id)
    if not _booking_has_scope_access(booking):
        return jsonify([]), 403

    products = []
    
    for item in booking.products:
        # Use custom unit price if set, otherwise product default price
        current_price = item.unit_price if item.unit_price is not None else item.product.price
        
        products.append({
            'product_id': item.product_id,
            'booking_id': item.booking_id,
            'quantity': item.quantity,
            'product_name': item.product.name_ar,
            'price': current_price,
            'original_price': item.product.price,
            'stock': item.product.stock_quantity
        })
    
    return jsonify(products)
        
@bp.route('/api/customer-vehicles/<int:customer_id>')
def get_customer_vehicles(customer_id):
    from app.models import Vehicle
    vehicles = Vehicle.query.filter_by(user_id=customer_id).all()
    v_list = []
    for v in vehicles:
        v_list.append({
            'id': v.id,
            'brand': v.brand,
            'plate_number': v.plate_number,
            'size_id': v.vehicle_size_id,
            'size_name': v.size.name_ar if v.size else ''
        })
    return jsonify(v_list)

@bp.route('/api/customers/search')
def search_customers():
    query_text = request.args.get('q', '').strip()
    if len(query_text) < 2:
        return jsonify([])

    pattern = f'%{query_text}%'
    customers = User.query.filter(
        User.role == 'customer',
        or_(
            User.username.ilike(pattern),
            User.phone.ilike(pattern),
            User.email.ilike(pattern)
        )
    ).order_by(User.username.asc()).limit(20).all()

    return jsonify([{
        'id': customer.id,
        'name': customer.username,
        'phone': customer.phone or '',
        'email': customer.email or ''
    } for customer in customers])

def _employee_available_slots(employee_id, date_obj, service_id=None, exclude_booking_id=None):
    """Return valid 30-minute start times for an employee."""
    from datetime import timedelta
    from app.utils.timezone import get_saudi_time

    employee = User.query.get(employee_id)
    if not employee or employee.role != 'employee' or _is_employee_on_break(employee, date_obj):
        return []

    booking_duration = 90
    if service_id:
        service = Service.query.get(service_id)
        if service and service.duration:
            booking_duration = service.duration

    schedules = EmployeeSchedule.query.filter_by(
        employee_id=employee_id,
        day_of_week=date_obj.weekday(),
        is_active=True
    ).all()
    if not schedules:
        return []

    bookings_query = Booking.query.filter(
        Booking.employee_id == employee_id,
        Booking.date == date_obj,
        Booking.status != 'cancelled'
    )
    if exclude_booking_id:
        bookings_query = bookings_query.filter(Booking.id != exclude_booking_id)

    blocked_ranges = []
    for existing_booking in bookings_query.all():
        duration = existing_booking.service.duration if (
            existing_booking.service and existing_booking.service.duration
        ) else 90
        booking_start = datetime.combine(date_obj, existing_booking.time)
        blocked_ranges.append((booking_start, booking_start + timedelta(minutes=duration)))

    now = get_saudi_time()
    is_today = date_obj == now.date()
    slots = set()
    for schedule in schedules:
        current = datetime.combine(date_obj, schedule.start_time)
        if schedule.end_time < schedule.start_time:
            end = datetime.combine(date_obj + timedelta(days=1), schedule.end_time)
        else:
            end = datetime.combine(date_obj, schedule.end_time)

        while current < end:
            slot_end = current + timedelta(minutes=booking_duration)
            if slot_end > end:
                break
            if is_today and current <= (now + timedelta(minutes=30)).replace(tzinfo=None):
                current += timedelta(minutes=30)
                continue
            overlaps = any(current < blocked_end and slot_end > blocked_start
                           for blocked_start, blocked_end in blocked_ranges)
            if not overlaps and not employee_break_overlaps(employee, current, slot_end):
                slots.add(current.strftime('%H:%M'))
            current += timedelta(minutes=30)

    return sorted(slots)

@bp.route('/api/available-slots/<int:employee_id>/<date>')
def get_available_slots(employee_id, date):
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify([])

    exclude_booking_id = request.args.get('exclude_booking_id', type=int)
    if exclude_booking_id:
        booking = Booking.query.get(exclude_booking_id)
        if not booking or not _booking_has_scope_access(booking) or booking.employee_id != employee_id:
            return jsonify([]), 403

    return jsonify(_employee_available_slots(
        employee_id,
        date_obj,
        request.args.get('service_id', type=int),
        exclude_booking_id=exclude_booking_id
    ))

@bp.route('/api/area-available-slots/<int:neighborhood_id>/<date>')
def get_area_available_slots(neighborhood_id, date):
    """Get all available time slots from all employees in a neighborhood"""
    from datetime import datetime, timedelta
    from app.models import employee_neighborhoods
    from app.utils.timezone import get_saudi_time
    
    # Get service duration if provided
    service_id = request.args.get('service_id')
    booking_duration = 90 # Default
    if service_id:
        from app.models import Service
        service = Service.query.get(service_id)
        if service and service.duration:
            booking_duration = service.duration

    date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    day_of_week = date_obj.weekday()
    
    # Get current Saudi time for "today" check
    now = get_saudi_time()
    is_today = date_obj == now.date()
    
    # Get all employees in this neighborhood
    employees = User.query.join(employee_neighborhoods).filter(
        employee_neighborhoods.c.neighborhood_id == neighborhood_id,
        User.role == 'employee'
    ).all()
    employees = [emp for emp in employees if not _is_employee_on_break(emp, date_obj)]
    
    # Collect all available slots from all employees
    all_slots = set()
    for emp in employees:
        schedules = EmployeeSchedule.query.filter_by(
            employee_id=emp.id,
            day_of_week=day_of_week,
            is_active=True
        ).all()
        
        # Get bookings for this employee, excluding cancelled ones
        bookings = Booking.query.filter(
            Booking.employee_id == emp.id,
            Booking.date == date_obj,
            Booking.status != 'cancelled'
        ).all()
        
        # Build blocked ranges
        blocked_ranges = []
        for booking in bookings:
            duration = booking.service.duration if (booking.service and booking.service.duration) else 90
            booking_start = datetime.combine(date_obj, booking.time)
            booking_end = booking_start + timedelta(minutes=duration)
            blocked_ranges.append((booking_start, booking_end))
        
        for schedule in schedules:
            current = datetime.combine(date_obj, schedule.start_time)
            
            # Handle night shift
            if schedule.end_time < schedule.start_time:
                end = datetime.combine(date_obj + timedelta(days=1), schedule.end_time)
            else:
                end = datetime.combine(date_obj, schedule.end_time)
            
            while current < end:
                slot_end = current + timedelta(minutes=booking_duration)
                
                # Skip if slot would extend beyond working hours
                if slot_end > end:
                    break
                
                time_str = current.strftime('%H:%M')
                
                # Skip if it's today and the time has passed (plus 30 min buffer)
                if is_today and current <= (now + timedelta(minutes=30)).replace(tzinfo=None):
                    current = current + timedelta(minutes=30)
                    continue
                
                # Check if this slot overlaps with any blocked range
                slot_blocked = False
                for blocked_start, blocked_end in blocked_ranges:
                    if current < blocked_end and slot_end > blocked_start:
                        slot_blocked = True
                        break
                
                if employee_break_overlaps(emp, current, slot_end):
                    slot_blocked = True

                if not slot_blocked:
                    all_slots.add(time_str)
                
                current = current + timedelta(minutes=30)
    
    return jsonify(sorted(list(all_slots)))

def auto_assign_employee(neighborhood_id, date, time_str):
    """Automatically assign an available employee from the neighborhood"""
    from datetime import datetime, time as dt_time
    from app.models import employee_neighborhoods
    
    date_obj = datetime.strptime(date, '%Y-%m-%d').date() if isinstance(date, str) else date
    day_of_week = date_obj.weekday()
    
    # Convert time string to time object for comparison
    hour, minute = map(int, time_str.split(':'))
    time_obj = dt_time(hour, minute)
    
    # Get all employees in this neighborhood
    employees = User.query.join(employee_neighborhoods).filter(
        employee_neighborhoods.c.neighborhood_id == neighborhood_id,
        User.role == 'employee'
    ).all()
    employees = [emp for emp in employees if not _is_employee_on_break(emp, date_obj, time_obj)]
    
    available_employees = []
    for emp in employees:
        # Check if employee has schedule for this day
        schedule = EmployeeSchedule.query.filter_by(
            employee_id=emp.id,
            day_of_week=day_of_week,
            is_active=True
        ).first()
        
        if not schedule:
            continue
            
        # Check if time is within working hours
        if not (schedule.start_time <= time_obj < schedule.end_time):
            continue
        
        # Check if not already booked
        existing_booking = Booking.query.filter_by(
            employee_id=emp.id,
            date=date_obj,
            time=time_str
        ).first()
        
        if not existing_booking:
            # Count today's bookings for load balancing
            bookings_count = Booking.query.filter_by(
                employee_id=emp.id,
                date=date_obj
            ).count()
            available_employees.append((emp, bookings_count))
    
    if not available_employees:
        return None
    
    # Sort by bookings count (load balancing) and return employee with least bookings
    available_employees.sort(key=lambda x: x[1])
    return available_employees[0][0].id

@bp.route('/bookings/<int:id>/reassign', methods=['POST'])
def reassign_booking(id):
    """Reassign booking to a different employee in the same neighborhood"""
    from datetime import datetime, time as dt_time
    
    booking = Booking.query.get_or_404(id)
    new_employee_id = request.form.get('employee_id')
    new_time_str = request.form.get('time')  # Optional - can change time too
    
    if not new_employee_id:
        flash('يجب اختيار موظف')
        return redirect(url_for('admin.bookings'))

    new_employee = User.query.get(int(new_employee_id))
    booking_start = datetime.combine(booking.date, booking.time)
    booking_end = booking_start + timedelta(minutes=booking.total_duration)
    if (not new_employee or new_employee.role != 'employee'
            or employee_break_overlaps(new_employee, booking_start, booking_end)):
        flash('لا يمكن إسناد الحجز لموظف في وضع الراحة')
        return redirect(url_for('admin.bookings'))
    
    # If time is being changed, convert and validate
    if new_time_str:
        hour, minute = map(int, new_time_str.split(':'))
        new_time = dt_time(hour, minute)
        
        # Check if new employee is available at new time
        existing = Booking.query.filter_by(
            employee_id=int(new_employee_id),
            date=booking.date,
            time=new_time_str
        ).first()
        
        if existing:
            flash('الموظف محجوز في هذا الوقت')
            return redirect(url_for('admin.bookings'))
        
        booking.time = new_time
    else:
        # Check availability at current time
        existing = Booking.query.filter_by(
            employee_id=int(new_employee_id),
            date=booking.date,
            time=booking.time.strftime('%H:%M')
        ).first()
        
        if existing:
            flash('الموظف محجوز في هذا الوقت')
            return redirect(url_for('admin.bookings'))
    
    booking.employee_id = int(new_employee_id)
    db.session.commit()
    flash('تم إعادة إسناد الحجز بنجاح')
    return redirect(url_for('admin.bookings'))

@bp.route('/bookings/<int:id>/advance-status', methods=['POST'])
def advance_booking_status(id):
    from app.utils.timezone import get_saudi_time

    booking = Booking.query.get_or_404(id)
    next_status = BOOKING_NEXT_STATUS.get(booking.status)

    if not next_status:
        flash('لا توجد خطوة تالية لهذا الحجز', 'warning')
        return redirect(request.referrer or url_for('admin.bookings'))

    if not booking.employee_id:
        flash('يجب إسناد الحجز لموظف قبل تغيير خطوات الخدمة', 'error')
        return redirect(request.referrer or url_for('admin.bookings'))

    now = get_saudi_time().replace(tzinfo=None)
    booking.status = next_status

    if next_status == 'in_progress' and not booking.started_at:
        booking.started_at = now

    completion_message = None
    if next_status == 'completed':
        if not booking.started_at:
            booking.started_at = now
        booking.completed_at = now
        completion_message = _apply_booking_completion_effects(booking)

    db.session.commit()

    if next_status in ['en_route', 'arrived', 'in_progress']:
        try:
            _send_booking_status_notification(booking, next_status)
        except Exception as e:
            print(f"Error sending status notification: {e}")
    elif next_status == 'completed':
        _send_booking_rating_request(booking)

    label = BOOKING_STATUS_LABELS.get(next_status, next_status)
    flash(completion_message or f'تم نقل الحجز إلى خطوة: {label}', 'success')
    return redirect(request.referrer or url_for('admin.bookings'))

@bp.route('/bookings/<int:id>/cancel', methods=['POST'])
def cancel_booking(id):
    from app.utils.timezone import get_saudi_time
    
    booking = Booking.query.get_or_404(id)
    
    if booking.status != 'cancelled':
        # Return products to stock
        for bp in booking.products:
            if bp.product:
                bp.product.stock_quantity += bp.quantity
                
        booking.status = 'cancelled'
        booking.cancelled_at = get_saudi_time().replace(tzinfo=None)
        db.session.commit()
        flash('تم إلغاء الحجز وإعادة المنتجات للمخزون')
    
    return redirect(url_for('admin.bookings'))

@bp.route('/bookings/<int:id>/delete', methods=['POST'])
def delete_booking(id):
    booking = Booking.query.get_or_404(id)
    
    # Return products to stock if not already cancelled (assuming cancellation returns stock)
    if booking.status != 'cancelled':
        for bp in booking.products:
            if bp.product:
                bp.product.stock_quantity += bp.quantity
                
    db.session.delete(booking)
    db.session.commit()
    flash('تم حذف الحجز نهائياً')
    return redirect(url_for('admin.bookings'))

# --- Reports ---
@bp.route('/reports')
def reports():
    from sqlalchemy import func, and_
    from datetime import datetime, timedelta
    
    # Get query parameters
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    city_id = request.args.get('city_id', type=int)
    neighborhood_id = request.args.get('neighborhood_id', type=int)
    
    # Set default date range (last 30 days if not specified)
    if not from_date_str:
        from_date = (datetime.now() - timedelta(days=30)).date()
    else:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    
    if not to_date_str:
        to_date = datetime.now().date()
    else:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

    supervisor_neighborhood_ids = _supervisor_neighborhood_ids() if current_user.role == 'supervisor' else None
    supervisor_city_ids = set()
    if supervisor_neighborhood_ids is not None and supervisor_neighborhood_ids:
        supervisor_city_ids = {
            n.city_id for n in Neighborhood.query.filter(Neighborhood.id.in_(supervisor_neighborhood_ids)).all()
        }
        if neighborhood_id and neighborhood_id not in supervisor_neighborhood_ids:
            neighborhood_id = None
        if city_id and city_id not in supervisor_city_ids:
            city_id = None
    
    # Base queries with date filter
    bookings_query = Booking.query.join(Neighborhood).filter(
        Booking.date >= from_date,
        Booking.date <= to_date
    )
    
    completed_bookings_query = Booking.query.join(Neighborhood).filter(
        Booking.date >= from_date,
        Booking.date <= to_date + timedelta(days=1),
        Booking.status == 'completed'
    )
    
    # Filter by city or neighborhood if specified
    if city_id:
        bookings_query = bookings_query.filter(Neighborhood.city_id == city_id)
        completed_bookings_query = completed_bookings_query.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        bookings_query = bookings_query.filter(Booking.neighborhood_id == neighborhood_id)
        completed_bookings_query = completed_bookings_query.filter(Booking.neighborhood_id == neighborhood_id)
    
    customers_query = User.query.filter_by(role='customer')
    
    subscriptions_query = Subscription.query.filter(
        Subscription.start_date >= from_date,
        Subscription.start_date <= to_date,
        Subscription.status == 'active'
    )
    
    # Payment Method Stats
    cash_bookings = Booking.query.join(Neighborhood).filter(
        Booking.date >= from_date,
        Booking.date <= to_date + timedelta(days=1),
        Booking.status == 'completed',
        Booking.payment_method == 'cash'
    )
    
    card_bookings = Booking.query.join(Neighborhood).filter(
        Booking.date >= from_date,
        Booking.date <= to_date + timedelta(days=1),
        Booking.status == 'completed',
        Booking.payment_method == 'card'
    )
    
    if city_id:
        cash_bookings = cash_bookings.filter(Neighborhood.city_id == city_id)
        card_bookings = card_bookings.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        cash_bookings = cash_bookings.filter(Booking.neighborhood_id == neighborhood_id)
        card_bookings = card_bookings.filter(Booking.neighborhood_id == neighborhood_id)

    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = supervisor_neighborhood_ids or []
        if supervisor_neighborhood_ids:
            cash_bookings = cash_bookings.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
            card_bookings = card_bookings.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            cash_bookings = cash_bookings.filter(Booking.id == -1)
            card_bookings = card_bookings.filter(Booking.id == -1)
        
    cash_bookings_list = _filter_bookings_by_work_date(
        cash_bookings.all(), from_date, to_date
    )
    card_bookings_list = _filter_bookings_by_work_date(
        card_bookings.all(), from_date, to_date
    )
    cash_count = len(cash_bookings_list)
    card_count = len(card_bookings_list)
    
    # Calculate totals for cash/card
    cash_total = 0
    for b in cash_bookings_list:
        if not b.service:
            continue
            
        # Check if subscription or free wash -> Service is 0
        if b.subscription_id or b.used_free_wash:
            # Calculate products only (using unit_price if set)
            p_total = sum((bp.unit_price if bp.unit_price is not None else bp.product.price) * bp.quantity for bp in b.products)
            price = p_total
        else:
            # Standard booking
            # Use custom service price if set
            service_price = b.custom_service_price if b.custom_service_price is not None else b.service.price
            price = service_price + (b.vehicle_size_price or 0)
            
            # Add products
            p_total = sum((bp.unit_price if bp.unit_price is not None else (bp.product.price if (bp.product and bp.product.price is not None) else 0)) * bp.quantity for bp in b.products)
            price += p_total
            
            # Apply discount
            if b.discount_code:
                if b.discount_code.discount_type == 'percentage':
                    disc = (service_price + (b.vehicle_size_price or 0)) * ((b.discount_code.value or 0) / 100)
                    price -= disc
                else:
                    price -= b.discount_code.value
                price = max(0, price)
        
        cash_total += price

    card_total = 0
    for b in card_bookings_list:
        if not b.service:
            continue
            
        # Check if subscription or free wash -> Service is 0
        if b.subscription_id or b.used_free_wash:
            # Calculate products only
            p_total = sum((bp.unit_price if bp.unit_price is not None else (bp.product.price if (bp.product and bp.product.price is not None) else 0)) * bp.quantity for bp in b.products)
            price = p_total
        else:
            # Standard booking
            service_price = b.custom_service_price if b.custom_service_price is not None else b.service.price
            price = service_price + (b.vehicle_size_price or 0)
            
            # Add products
            p_total = sum((bp.unit_price if bp.unit_price is not None else (bp.product.price if (bp.product and bp.product.price is not None) else 0)) * bp.quantity for bp in b.products)
            price += p_total
            
            # Apply discount
            if b.discount_code:
                if b.discount_code.discount_type == 'percentage':
                    disc = (service_price + (b.vehicle_size_price or 0)) * ((b.discount_code.value or 0) / 100)
                    price -= disc
                else:
                    price -= b.discount_code.value
                price = max(0, price)
        
        card_total += price

    # Filter for supervisor
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = supervisor_neighborhood_ids or []
        
        if supervisor_neighborhood_ids:
            bookings_query = bookings_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
            completed_bookings_query = completed_bookings_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
            
            # Filter customers who have bookings in supervisor's area
            # Fix AmbiguousForeignKeysError by specifying join condition
            customers_query = customers_query.join(Booking, User.id == Booking.customer_id).filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids)).distinct()
            
            subscriptions_query = subscriptions_query.filter(Subscription.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            # No scope assigned
            bookings_query = bookings_query.filter_by(id=-1)
            completed_bookings_query = completed_bookings_query.filter_by(id=-1)
            customers_query = customers_query.filter_by(id=-1)
            subscriptions_query = subscriptions_query.filter_by(id=-1)

    total_bookings = bookings_query.count()
    completed_bookings_list = _filter_bookings_by_work_date(
        completed_bookings_query.all(), from_date, to_date
    )
    completed_bookings = len(completed_bookings_list)
    total_customers = customers_query.count()
    active_subscriptions = subscriptions_query.count()
    
    # Revenue calculations with accurate pricing
    service_revenue = 0
    product_revenue = 0
    
    for booking in completed_bookings_list:
        # Calculate service price after discount/free wash
        service_price = booking.custom_service_price if booking.custom_service_price is not None else (booking.service.price if booking.service else 0)
        discount_amount = 0
        
        # Check if subscription or free wash -> Service is 0
        if booking.subscription_id or booking.used_free_wash:
            service_price = 0
        # Check if discount code was applied
        elif booking.discount_code:
            if booking.discount_code.discount_type == 'percentage':
                discount_amount = (service_price + (booking.vehicle_size_price or 0)) * ((booking.discount_code.value or 0) / 100)
            else:
                discount_amount = booking.discount_code.value
        
        # Add service revenue (including vehicle size price)
        final_service_price = max(0, service_price - discount_amount + (booking.vehicle_size_price or 0))
        if booking.subscription_id or booking.used_free_wash:
             final_service_price = 0
             
        service_revenue += final_service_price
        
        # Add product revenue
        for bp in booking.products:
            product_revenue += ((bp.unit_price if bp.unit_price is not None else (bp.product.price if (bp.product and bp.product.price is not None) else 0)) * bp.quantity)
    
    # Subscription revenue (only active subscriptions created in date range)
    sub_rev_query = db.session.query(func.sum(SubscriptionPackage.price))\
        .join(Subscription)\
        .join(Neighborhood, Subscription.neighborhood_id == Neighborhood.id)\
        .filter(
            Subscription.start_date >= from_date,
            Subscription.start_date <= to_date,
            Subscription.status == 'active'
        )
        
    if city_id:
        sub_rev_query = sub_rev_query.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        sub_rev_query = sub_rev_query.filter(Subscription.neighborhood_id == neighborhood_id)
        
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            sub_rev_query = sub_rev_query.filter(Subscription.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            sub_rev_query = sub_rev_query.filter(Subscription.id == -1)
        
    subscription_revenue = sub_rev_query.scalar() or 0
    
    # Total revenue
    total_revenue = service_revenue + product_revenue + subscription_revenue
    
    # Top services (in date range)
    top_services_query = db.session.query(
        Service.name_ar,
        func.count(Booking.id).label('count')
    ).join(Booking)\
    .join(Neighborhood, Booking.neighborhood_id == Neighborhood.id)\
    .filter(
        Booking.date >= from_date,
        Booking.date <= to_date
    )
    
    if city_id:
        top_services_query = top_services_query.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        top_services_query = top_services_query.filter(Booking.neighborhood_id == neighborhood_id)
    
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            top_services_query = top_services_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            top_services_query = top_services_query.filter(Booking.id == -1)
        
    top_services = top_services_query.group_by(Service.id)\
    .order_by(func.count(Booking.id).desc())\
    .limit(5).all()
    
    # Employee performance uses the logical shift date, so work completed
    # after midnight remains attached to the shift that started the day before.
    employee_bookings_query = Booking.query.join(
        User, User.id == Booking.employee_id
    ).join(
        Neighborhood, Booking.neighborhood_id == Neighborhood.id
    ).filter(
        User.role == 'employee',
        Booking.date >= from_date,
        Booking.date <= to_date + timedelta(days=1)
    )
    
    if city_id:
        employee_bookings_query = employee_bookings_query.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        employee_bookings_query = employee_bookings_query.filter(Booking.neighborhood_id == neighborhood_id)
    
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            employee_bookings_query = employee_bookings_query.filter(Booking.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            employee_bookings_query = employee_bookings_query.filter(Booking.id == -1)

    employee_bookings = _filter_bookings_by_work_date(
        employee_bookings_query.all(), from_date, to_date
    )
    employee_totals = {}
    for booking in employee_bookings:
        row = employee_totals.setdefault(
            booking.employee_id,
            {'username': booking.employee.username, 'total': 0, 'completed': 0}
        )
        row['total'] += 1
        if booking.status == 'completed':
            row['completed'] += 1
    employee_stats = sorted(
        employee_totals.values(), key=lambda row: row['completed'], reverse=True
    )
    
    # City performance (in date range)
    city_performance = {} # {city_name: {'count': 0, 'revenue': 0}}
    for booking in completed_bookings_list:
        city_name = booking.neighborhood.city.name_ar if (booking.neighborhood and booking.neighborhood.city) else 'غير محدد'
        if city_name not in city_performance:
            city_performance[city_name] = {'count': 0, 'revenue': 0}
        
        # Calculate booking revenue (service + products - discount)
        # Service
        s_price = booking.custom_service_price if booking.custom_service_price is not None else (booking.service.price if booking.service else 0)
        if booking.subscription_id or booking.used_free_wash:
            s_price = 0
        else:
            d_amt = 0
            if booking.discount_code:
                if booking.discount_code.discount_type == 'percentage':
                    d_amt = (s_price + (booking.vehicle_size_price or 0)) * ((booking.discount_code.value or 0) / 100)
                else:
                    d_amt = booking.discount_code.value
            s_price = max(0, s_price - d_amt + (booking.vehicle_size_price or 0))
        
        # Products
        p_price = sum([(bp.unit_price if bp.unit_price is not None else (bp.product.price if (bp.product and bp.product.price is not None) else 0)) * bp.quantity for bp in booking.products])
        
        city_performance[city_name]['count'] += 1
        city_performance[city_name]['revenue'] += (s_price + p_price)

    # Add subscription revenue to city performance
    from app.models import City
    sub_city_rev_query = db.session.query(
        City.name_ar,
        func.sum(SubscriptionPackage.price)
    ).join(Neighborhood, Neighborhood.city_id == City.id)\
    .join(Subscription, Subscription.neighborhood_id == Neighborhood.id)\
    .join(SubscriptionPackage)\
    .filter(
        Subscription.start_date >= from_date,
        Subscription.start_date <= to_date,
        Subscription.status == 'active'
    )
    
    if city_id:
        sub_city_rev_query = sub_city_rev_query.filter(Neighborhood.city_id == city_id)
        
    if neighborhood_id:
        sub_city_rev_query = sub_city_rev_query.filter(Subscription.neighborhood_id == neighborhood_id)
        
    if current_user.role == 'supervisor':
        if supervisor_neighborhood_ids:
            sub_city_rev_query = sub_city_rev_query.filter(Subscription.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            sub_city_rev_query = sub_city_rev_query.filter(Subscription.id == -1)
            
    sub_city_revs = sub_city_rev_query.group_by(City.id).all()
    
    for c_name, c_rev in sub_city_revs:
        if c_name not in city_performance:
            city_performance[c_name] = {'count': 0, 'revenue': 0}
        city_performance[c_name]['revenue'] += (c_rev or 0)

    # Fetch cities and neighborhoods for the filters
    if current_user.role == 'supervisor':
        cities = City.query.filter(City.is_active == True, City.id.in_(supervisor_city_ids)).all() if supervisor_city_ids else []
    else:
        cities = City.query.filter_by(is_active=True).all()
    neighborhoods = []
    if city_id:
        neighborhoods_query = Neighborhood.query.filter_by(city_id=city_id, is_active=True)
        if current_user.role == 'supervisor':
            neighborhoods_query = neighborhoods_query.filter(Neighborhood.id.in_(supervisor_neighborhood_ids)) if supervisor_neighborhood_ids else neighborhoods_query.filter(Neighborhood.id == -1)
        neighborhoods = neighborhoods_query.all()

    return render_template('admin/reports.html', 
                           total_bookings=total_bookings,
                           completed_bookings=completed_bookings,
                           total_customers=total_customers,
                           active_subscriptions=active_subscriptions,
                           service_revenue=service_revenue,
                           product_revenue=product_revenue,
                           subscription_revenue=subscription_revenue,
                           total_revenue=total_revenue,
                           top_services=top_services,
                           employee_stats=employee_stats,
                           city_performance=city_performance,
                           cash_count=cash_count,
                           cash_total=cash_total,
                           card_count=card_count,
                           card_total=card_total,
                           from_date=from_date.strftime('%Y-%m-%d'),
                           to_date=to_date.strftime('%Y-%m-%d'),
                           city_id=city_id,
                           neighborhood_id=neighborhood_id,
                           cities=cities,
                           neighborhoods=neighborhoods)

@bp.route('/management-reports')
def management_reports():
    """Decision-focused operational reports; intentionally excludes profitability."""
    today = datetime.now().date()
    try:
        from_date = datetime.strptime(request.args.get('from_date', ''), '%Y-%m-%d').date()
    except (TypeError, ValueError):
        from_date = today - timedelta(days=29)
    try:
        to_date = datetime.strptime(request.args.get('to_date', ''), '%Y-%m-%d').date()
    except (TypeError, ValueError):
        to_date = today
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    city_id = request.args.get('city_id', type=int)
    neighborhood_id = request.args.get('neighborhood_id', type=int)
    scoped_ids = _supervisor_neighborhood_ids()
    allowed_city_ids = None
    if scoped_ids is not None:
        allowed_city_ids = {n.city_id for n in Neighborhood.query.filter(Neighborhood.id.in_(scoped_ids)).all()} if scoped_ids else set()
        if neighborhood_id not in (scoped_ids or []):
            neighborhood_id = None
        if city_id not in (allowed_city_ids or set()):
            city_id = None

    def scoped_booking_query(start, end):
        query = Booking.query.outerjoin(Neighborhood).filter(Booking.date.between(start, end))
        if city_id:
            query = query.filter(Neighborhood.city_id == city_id)
        if neighborhood_id:
            query = query.filter(Booking.neighborhood_id == neighborhood_id)
        if scoped_ids is not None:
            query = query.filter(Booking.neighborhood_id.in_(scoped_ids)) if scoped_ids else query.filter(Booking.id == -1)
        return query

    bookings = scoped_booking_query(from_date, to_date).all()
    total = len(bookings)
    completed = sum(b.status == 'completed' for b in bookings)
    cancelled = sum(b.status == 'cancelled' for b in bookings)
    completion_rate = round(completed * 100 / total, 1) if total else 0
    cancellation_rate = round(cancelled * 100 / total, 1) if total else 0
    ratings = [b.rating for b in bookings if b.rating is not None]
    average_rating = round(sum(ratings) / len(ratings), 1) if ratings else 0

    period_days = (to_date - from_date).days + 1
    previous_end = from_date - timedelta(days=1)
    previous_start = previous_end - timedelta(days=period_days - 1)
    previous_total = scoped_booking_query(previous_start, previous_end).count()
    booking_growth = round((total - previous_total) * 100 / previous_total, 1) if previous_total else (100 if total else 0)

    customer_counts = {}
    for booking in bookings:
        if booking.customer_id:
            customer_counts[booking.customer_id] = customer_counts.get(booking.customer_id, 0) + 1
    new_customers = 0
    for customer_id in customer_counts:
        first_date = db.session.query(func.min(Booking.date)).filter(Booking.customer_id == customer_id).scalar()
        if first_date and from_date <= first_date <= to_date:
            new_customers += 1
    returning_customers = sum(count > 1 for count in customer_counts.values())

    employee_map, service_map, city_map, hour_counts = {}, {}, {}, {hour: 0 for hour in range(24)}
    cancellation_reasons = {}
    discount_map = {}
    durations = []
    for booking in bookings:
        if booking.time:
            hour_counts[booking.time.hour] += 1
        if booking.employee:
            row = employee_map.setdefault(booking.employee_id, {'name': booking.employee.username, 'total': 0, 'completed': 0, 'rating_sum': 0, 'rating_count': 0})
            row['total'] += 1
            row['completed'] += booking.status == 'completed'
            if booking.rating is not None:
                row['rating_sum'] += booking.rating
                row['rating_count'] += 1
        if booking.service:
            row = service_map.setdefault(booking.service_id, {'name': booking.service.name_ar, 'count': 0})
            row['count'] += 1
        if booking.neighborhood and booking.neighborhood.city:
            row = city_map.setdefault(booking.neighborhood.city_id, {'name': booking.neighborhood.city.name_ar, 'total': 0, 'completed': 0, 'cancelled': 0})
            row['total'] += 1
            row['completed'] += booking.status == 'completed'
            row['cancelled'] += booking.status == 'cancelled'
        if booking.status == 'cancelled':
            reason = (booking.cancellation_reason or 'غير محدد').strip() or 'غير محدد'
            cancellation_reasons[reason] = cancellation_reasons.get(reason, 0) + 1
        if booking.discount_code:
            code = booking.discount_code.code
            discount_map[code] = discount_map.get(code, 0) + 1
        if booking.started_at and booking.completed_at and booking.completed_at >= booking.started_at:
            durations.append((booking.completed_at - booking.started_at).total_seconds() / 60)

    employee_stats = []
    for row in employee_map.values():
        row['rate'] = round(row['completed'] * 100 / row['total'], 1)
        row['rating'] = round(row['rating_sum'] / row['rating_count'], 1) if row['rating_count'] else 0
        employee_stats.append(row)
    employee_stats.sort(key=lambda row: (row['completed'], row['rating']), reverse=True)
    service_stats = sorted(service_map.values(), key=lambda row: row['count'], reverse=True)[:10]
    city_stats = sorted(city_map.values(), key=lambda row: row['total'], reverse=True)
    for row in city_stats:
        row['completion_rate'] = round(row['completed'] * 100 / row['total'], 1) if row['total'] else 0

    subscription_query = Subscription.query.filter(Subscription.start_date.between(from_date, to_date))
    if city_id:
        subscription_query = subscription_query.join(Neighborhood).filter(Neighborhood.city_id == city_id)
    if neighborhood_id:
        subscription_query = subscription_query.filter(Subscription.neighborhood_id == neighborhood_id)
    if scoped_ids is not None:
        subscription_query = subscription_query.filter(Subscription.neighborhood_id.in_(scoped_ids)) if scoped_ids else subscription_query.filter(Subscription.id == -1)
    subscriptions = subscription_query.all()
    active_subscriptions = sum(s.status == 'active' for s in subscriptions)
    expiring_subscriptions = Subscription.query.filter(Subscription.status == 'active', Subscription.end_date.between(today, today + timedelta(days=7)))
    if scoped_ids is not None:
        expiring_subscriptions = expiring_subscriptions.filter(Subscription.neighborhood_id.in_(scoped_ids)) if scoped_ids else expiring_subscriptions.filter(Subscription.id == -1)
    expiring_count = expiring_subscriptions.count()

    alerts = []
    if cancellation_rate >= 15:
        alerts.append(f'نسبة الإلغاء مرتفعة ({cancellation_rate}%)')
    if booking_growth <= -15:
        alerts.append(f'انخفضت الطلبات {abs(booking_growth)}% عن الفترة السابقة')
    if average_rating and average_rating < 4:
        alerts.append(f'متوسط التقييم منخفض ({average_rating} من 5)')
    if expiring_count:
        alerts.append(f'{expiring_count} اشتراك سينتهي خلال 7 أيام')

    cities_query = City.query.filter_by(is_active=True)
    if allowed_city_ids is not None:
        cities_query = cities_query.filter(City.id.in_(allowed_city_ids)) if allowed_city_ids else cities_query.filter(City.id == -1)
    cities = cities_query.order_by(City.name_ar).all()
    neighborhoods_query = Neighborhood.query.filter_by(is_active=True)
    if city_id:
        neighborhoods_query = neighborhoods_query.filter_by(city_id=city_id)
    if scoped_ids is not None:
        neighborhoods_query = neighborhoods_query.filter(Neighborhood.id.in_(scoped_ids)) if scoped_ids else neighborhoods_query.filter(Neighborhood.id == -1)

    return render_template('admin/management_reports.html',
        from_date=from_date, to_date=to_date, city_id=city_id, neighborhood_id=neighborhood_id,
        cities=cities, neighborhoods=neighborhoods_query.order_by(Neighborhood.name_ar).all(),
        total=total, completed=completed, cancelled=cancelled, completion_rate=completion_rate,
        cancellation_rate=cancellation_rate, booking_growth=booking_growth, average_rating=average_rating,
        unique_customers=len(customer_counts), new_customers=new_customers, returning_customers=returning_customers,
        average_duration=round(sum(durations) / len(durations)) if durations else 0,
        employee_stats=employee_stats[:10], service_stats=service_stats, city_stats=city_stats,
        peak_hours=sorted(hour_counts.items(), key=lambda item: item[1], reverse=True)[:6],
        cancellation_reasons=sorted(cancellation_reasons.items(), key=lambda item: item[1], reverse=True)[:6],
        discount_stats=sorted(discount_map.items(), key=lambda item: item[1], reverse=True)[:6],
        new_subscriptions=len(subscriptions), active_subscriptions=active_subscriptions,
        expiring_subscriptions=expiring_count, alerts=alerts)

# --- Settings (Loyalty, Admin Accounts, Backup) ---
@bp.route('/settings/loyalty', methods=['GET', 'POST'])
def loyalty_settings():
    settings = SiteSettings.get_settings()
    
    if request.method == 'POST':
        threshold = request.form.get('threshold', type=int)
        if threshold and threshold > 0:
            settings.loyalty_points_threshold = threshold
            db.session.commit()
            flash(f'تم تحديث عتبة الولاء إلى {threshold} نقطة')
        else:
            flash('الرجاء إدخال قيمة صحيحة', 'error')
    
    return render_template('admin/loyalty_settings.html', current_threshold=settings.loyalty_points_threshold)

@bp.route('/backup/export-json')
def backup_json():
    import json
    from flask import Response
    
    data = {
        'users': [{'id': u.id, 'username': u.username, 'role': u.role} for u in User.query.all()],
        'bookings': [{'id': b.id, 'status': b.status, 'date': str(b.date)} for b in Booking.query.all()],
    }
    
    return Response(
        json.dumps(data, ensure_ascii=False, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': 'attachment; filename=backup.json'}
    )

@bp.route('/settings', methods=['GET', 'POST'])
def settings():
    form = SiteSettingsForm()
    settings = SiteSettings.get_settings()
    
    if form.validate_on_submit():
        settings.site_name = form.site_name.data
        settings.primary_color = form.primary_color.data
        settings.accent_color = form.accent_color.data
        settings.whatsapp_number = form.whatsapp_number.data
        settings.facebook_url = form.facebook_url.data
        settings.twitter_url = form.twitter_url.data
        settings.instagram_url = form.instagram_url.data
        settings.tiktok_url = form.tiktok_url.data
        settings.mawthooq_url = form.mawthooq_url.data
        settings.terms_content = form.terms_content.data
        settings.booking_days_limit = form.booking_days_limit.data
        settings.subscription_days_limit = form.subscription_days_limit.data
        settings.referral_target_count = form.referral_target_count.data
        settings.maintenance_mode = form.maintenance_mode.data
        
        if form.logo.data:
            import os
            from werkzeug.utils import secure_filename
            from flask import current_app
            
            file = form.logo.data
            filename = secure_filename(file.filename)
            # Ensure filename is unique or standard
            filename = 'logo.png' # Force standard name for simplicity or keep original
            
            # Save to static/uploads or static/images
            upload_dir = os.path.join(current_app.root_path, 'static', 'images')
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir)
                
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)
            settings.logo_path = f'/static/images/{filename}'
            
        db.session.commit()
        flash('تم تحديث إعدادات الموقع بنجاح', 'success')
        return redirect(url_for('admin.settings'))
        
    elif request.method == 'GET':
        form.site_name.data = settings.site_name
        form.primary_color.data = settings.primary_color
        form.accent_color.data = settings.accent_color
        form.whatsapp_number.data = settings.whatsapp_number
        form.facebook_url.data = settings.facebook_url
        form.twitter_url.data = settings.twitter_url
        form.instagram_url.data = settings.instagram_url
        form.tiktok_url.data = settings.tiktok_url
        form.mawthooq_url.data = settings.mawthooq_url
        form.terms_content.data = settings.terms_content
        form.booking_days_limit.data = settings.booking_days_limit
        form.subscription_days_limit.data = settings.subscription_days_limit
        form.referral_target_count.data = settings.referral_target_count or 10
        form.maintenance_mode.data = bool(settings.maintenance_mode)

    return render_template('admin/settings.html', form=form, settings=settings)

@bp.route('/notifications/send', methods=['GET', 'POST'])
@login_required
def send_notification():
    form = NotificationForm()
    # Populate user choices
    users = User.query.filter(User.role == 'customer').all()
    choices = [(0, 'All Customers')] + [(u.id, f"{u.username} ({u.phone})") for u in users]
    form.user_id.choices = choices

    if form.validate_on_submit():
        title = form.title.data
        message = form.message.data
        recipient_id = form.user_id.data

        targets = []
        if recipient_id == 0:
            targets = users
        else:
            targets = [User.query.get(recipient_id)]

        count = 0
        for user in targets:
            if not user: continue
            
            # 1. Create DB Notification
            notif = Notification(user_id=user.id, title=title, message=message)
            db.session.add(notif)
            
            # 2. Send Web Push using the improved notification function
            from app.notifications import send_push_notification
            notification_data = {
                "title": title,
                "body": message,
                "icon": "/static/images/logo.png",
                "badge": "/static/images/logo.png",
                "url": "/notifications"
            }
            send_push_notification(user, notification_data)
            
            count += 1
        
        db.session.commit()
        flash(f'Notification sent to {count} users.', 'success')
        return redirect(url_for('admin.send_notification'))

    return render_template('admin/notifications.html', form=form)


# --- Discount Code Management ---
def _discount_location_choices():
    if current_user.role == 'admin':
        return (
            City.query.order_by(City.name_ar).all(),
            Neighborhood.query.join(City).order_by(City.name_ar, Neighborhood.name_ar).all()
        )
    city_ids = {city.id for city in current_user.supervisor_cities}
    neighborhood_ids = set(_supervisor_neighborhood_ids() or [])
    cities = City.query.filter(City.id.in_(city_ids)).order_by(City.name_ar).all() if city_ids else []
    neighborhoods = (Neighborhood.query.filter(Neighborhood.id.in_(neighborhood_ids))
                     .join(City).order_by(City.name_ar, Neighborhood.name_ar).all()
                     if neighborhood_ids else [])
    return cities, neighborhoods


def _get_scoped_discount_or_404(code_id):
    query = DiscountCode.query.filter_by(id=code_id)
    if current_user.role == 'supervisor':
        query = query.filter_by(created_by_id=current_user.id)
    return query.first_or_404()


def _parse_discount_location():
    scope_type = request.form.get('scope_type', 'global')
    city_id = request.form.get('city_id', type=int)
    neighborhood_id = request.form.get('neighborhood_id', type=int)
    cities, neighborhoods = _discount_location_choices()
    allowed_city_ids = {city.id for city in cities}
    allowed_neighborhood_ids = {neighborhood.id for neighborhood in neighborhoods}
    if current_user.role == 'supervisor' and scope_type == 'global':
        raise ValueError('يجب على المشرف تحديد مدينة أو حي ضمن نطاقه.')
    if scope_type == 'city':
        if city_id not in allowed_city_ids:
            raise ValueError('المدينة المحددة خارج نطاق صلاحياتك.')
        return city_id, None
    if scope_type == 'neighborhood':
        if neighborhood_id not in allowed_neighborhood_ids:
            raise ValueError('الحي المحدد خارج نطاق صلاحياتك.')
        return None, neighborhood_id
    if scope_type != 'global' or current_user.role != 'admin':
        raise ValueError('نطاق كود الخصم غير صالح.')
    return None, None


@bp.route('/discount_codes')
def discount_codes():
    page = request.args.get('page', 1, type=int)
    query = DiscountCode.query.filter(or_(DiscountCode.is_influencer == False, DiscountCode.is_influencer == None))
    if current_user.role == 'supervisor':
        query = query.filter(DiscountCode.created_by_id == current_user.id)
    pagination = query.order_by(DiscountCode.created_at.desc() if hasattr(DiscountCode, 'created_at') else DiscountCode.id.desc()).paginate(page=page, per_page=50, error_out=False)
    codes = pagination.items
    return render_template('admin/discount_codes.html', discount_codes=codes, pagination=pagination)


@bp.route('/abandoned-checkouts')
def abandoned_checkouts():
    expiry_cutoff = datetime.utcnow() - timedelta(days=ABANDONED_CHECKOUT_RETENTION_DAYS)
    CheckoutSession.query.filter(
        CheckoutSession.status == 'active',
        CheckoutSession.created_at < expiry_cutoff
    ).delete(synchronize_session=False)
    db.session.commit()

    cutoff = datetime.utcnow() - timedelta(minutes=30)
    query = CheckoutSession.query.filter(
        CheckoutSession.status == 'active',
        or_(
            CheckoutSession.last_activity_at <= cutoff,
            CheckoutSession.recovery_discount_code_id.isnot(None)
        )
    )
    if current_user.role == 'supervisor':
        allowed_city_ids = {city.id for city in current_user.supervisor_cities}
        allowed_neighborhood_ids = set(_supervisor_neighborhood_ids() or [])
        scope_filters = []
        if allowed_neighborhood_ids:
            scope_filters.append(CheckoutSession.neighborhood_id.in_(allowed_neighborhood_ids))
        if allowed_city_ids:
            scope_filters.append(CheckoutSession.city_id.in_(allowed_city_ids))
        if scope_filters:
            query = query.filter(or_(*scope_filters))
        else:
            query = query.filter(CheckoutSession.id == -1)

    search = request.args.get('q', '').strip()
    flow_type = request.args.get('flow_type', '').strip()
    if search:
        query = query.join(User, CheckoutSession.customer_id == User.id).filter(or_(
            User.username.ilike(f'%{search}%'),
            User.phone.ilike(f'%{search}%'),
            User.email.ilike(f'%{search}%')
        ))
    if flow_type:
        query = query.filter(CheckoutSession.flow_type == flow_type)

    page = request.args.get('page', 1, type=int)
    customer_query = query.with_entities(
        CheckoutSession.customer_id,
        func.max(CheckoutSession.last_activity_at).label('latest_activity')
    ).group_by(CheckoutSession.customer_id).order_by(
        func.max(CheckoutSession.last_activity_at).desc()
    )
    pagination = customer_query.paginate(
        page=page, per_page=50, error_out=False
    )
    customer_ids = [item.customer_id for item in pagination.items]
    used_recovery_discount_customer_ids = set()
    if customer_ids:
        used_recovery_discount_customer_ids = {
            customer_id for (customer_id,) in db.session.query(
                Booking.customer_id
            ).join(
                DiscountCode, Booking.discount_code_id == DiscountCode.id
            ).filter(
                Booking.customer_id.in_(customer_ids),
                Booking.status != 'cancelled',
                DiscountCode.code.like('BACK%')
            ).distinct().all()
        }
    grouped_checkouts = {customer_id: [] for customer_id in customer_ids}
    if customer_ids:
        checkouts = query.filter(
            CheckoutSession.customer_id.in_(customer_ids)
        ).order_by(
            CheckoutSession.last_activity_at.desc()
        ).all()
        for checkout in checkouts:
            grouped_checkouts.setdefault(checkout.customer_id, []).append(checkout)

    rows = []
    for customer_id in customer_ids:
        customer_checkouts = grouped_checkouts.get(customer_id, [])
        if not customer_checkouts:
            continue
        checkout = customer_checkouts[0]
        attempts = []
        for attempt in customer_checkouts:
            try:
                attempt_data = json.loads(attempt.form_data or '{}')
            except (TypeError, ValueError):
                attempt_data = {}
            attempts.append({'checkout': attempt, 'data': attempt_data})
        rows.append({
            'checkout': checkout,
            'data': attempts[0]['data'],
            'promo_code': checkout.recovery_discount_code,
            'whatsapp_phone': _normalize_whatsapp_phone(checkout.customer.phone),
            'attempts': attempts,
            'attempts_count': len(attempts),
            'used_recovery_discount': customer_id in used_recovery_discount_customer_ids
        })

    return render_template(
        'admin/abandoned_checkouts.html',
        rows=rows,
        pagination=pagination,
        search=search,
        selected_flow=flow_type,
        cutoff_minutes=30
    )


@bp.route('/abandoned-checkouts/<int:checkout_id>/create-discount', methods=['POST'])
def create_abandoned_checkout_discount(checkout_id):
    checkout = CheckoutSession.query.get_or_404(checkout_id)
    if checkout.status != 'active' or not _checkout_has_scope_access(checkout):
        abort(404)
    if checkout.recovery_discount_code:
        flash('يوجد كود خصم مرتبط بهذه السلة بالفعل. يمكنك تعديله أو حذفه.', 'info')
        return redirect(url_for('admin.abandoned_checkouts'))
    customer_last_activity = checkout.last_activity_at

    discount_type = request.form.get('discount_type', 'percentage').strip().lower()
    try:
        value = float(request.form.get('value', 0))
    except (TypeError, ValueError):
        value = 0
    if discount_type not in ('percentage', 'fixed') or value <= 0:
        flash('يرجى تحديد قيمة خصم صحيحة.', 'error')
        return redirect(url_for('admin.abandoned_checkouts'))
    if discount_type == 'percentage' and value > 100:
        flash('نسبة الخصم لا يمكن أن تتجاوز 100%.', 'error')
        return redirect(url_for('admin.abandoned_checkouts'))

    alphabet = string.ascii_uppercase + string.digits
    while True:
        code_text = 'BACK' + ''.join(secrets.choice(alphabet) for _ in range(6))
        if not DiscountCode.query.filter_by(code=code_text).first():
            break

    # Give the recovery code the narrowest location known for this cart.
    neighborhood_id = checkout.neighborhood_id
    city_id = None if neighborhood_id else checkout.city_id
    code = DiscountCode(
        code=code_text,
        discount_type=discount_type,
        value=value,
        valid_from=datetime.utcnow(),
        valid_until=datetime.utcnow() + timedelta(days=7),
        usage_limit=1,
        max_uses_per_customer=1,
        is_active=True,
        assigned_customer_id=checkout.customer_id,
        city_id=city_id,
        neighborhood_id=neighborhood_id,
        created_by_id=current_user.id
    )
    db.session.add(code)
    db.session.flush()
    checkout.recovery_discount_code_id = code.id
    db.session.flush()
    # Administrative actions must not make an abandoned cart look active again.
    CheckoutSession.query.filter_by(id=checkout.id).update(
        {'last_activity_at': customer_last_activity}, synchronize_session=False
    )
    db.session.commit()
    flash('تم إنشاء كود الخصم، وستتم إضافته تلقائيًا إلى رسالة واتساب.', 'success')
    return redirect(url_for('admin.abandoned_checkouts'))


@bp.route('/abandoned-checkouts/<int:checkout_id>/delete-discount', methods=['POST'])
def delete_abandoned_checkout_discount(checkout_id):
    checkout = CheckoutSession.query.get_or_404(checkout_id)
    if not _checkout_has_scope_access(checkout):
        abort(404)
    code = checkout.recovery_discount_code
    if code:
        customer_last_activity = checkout.last_activity_at
        checkout.recovery_discount_code_id = None
        db.session.flush()
        abandoned_cutoff = datetime.utcnow() - timedelta(minutes=31)
        CheckoutSession.query.filter_by(id=checkout.id).update(
            {'last_activity_at': min(customer_last_activity, abandoned_cutoff)},
            synchronize_session=False
        )
        db.session.delete(code)
        db.session.commit()
        flash('تم حذف كود الخصم من السلة.', 'success')
    return redirect(url_for('admin.abandoned_checkouts'))

@bp.route('/discount_codes/add', methods=['GET', 'POST'])
def add_discount_code():
    cities, neighborhoods = _discount_location_choices()
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        discount_type = request.form.get('discount_type', 'percentage').lower()
        value = float(request.form.get('value', 0))
        valid_until_str = request.form.get('valid_until')
        usage_limit = request.form.get('usage_limit')
        max_uses_per_customer = request.form.get('max_uses_per_customer')
        
        try:
            if not code or discount_type not in ('percentage', 'fixed') or value <= 0:
                raise ValueError('بيانات كود الخصم غير صالحة.')
            if discount_type == 'percentage' and value > 100:
                raise ValueError('نسبة الخصم لا يمكن أن تتجاوز 100%.')
            city_id, neighborhood_id = _parse_discount_location()
            valid_until = datetime.strptime(valid_until_str, '%Y-%m-%d')
        except (ValueError, TypeError) as exc:
            flash(str(exc), 'error')
            return render_template('admin/add_discount_code.html', cities=cities, neighborhoods=neighborhoods)
        
        new_code = DiscountCode(
            code=code,
            discount_type=discount_type,
            value=value,
            valid_until=valid_until,
            usage_limit=int(usage_limit) if usage_limit else None,
            max_uses_per_customer=int(max_uses_per_customer) if max_uses_per_customer else 1,
            city_id=city_id,
            neighborhood_id=neighborhood_id,
            created_by_id=current_user.id
        )
        
        try:
            db.session.add(new_code)
            db.session.commit()
            flash('تم إضافة كود الخصم بنجاح', 'success')
            return redirect(url_for('admin.discount_codes'))
        except:
            db.session.rollback()
            flash('حدث خطأ أثناء إضافة الكود. ربما الكود موجود مسبقاً.', 'error')
            
    return render_template('admin/add_discount_code.html', cities=cities, neighborhoods=neighborhoods)

@bp.route('/discount_codes/edit/<int:id>', methods=['GET', 'POST'])
def edit_discount_code(id):
    code = _get_scoped_discount_or_404(id)
    cities, neighborhoods = _discount_location_choices()
    
    if request.method == 'POST':
        code.code = request.form.get('code', '').strip().upper()
        discount_type = request.form.get('discount_type', 'percentage').lower()
        code.discount_type = discount_type
        code.value = float(request.form.get('value', 0))
        valid_until_str = request.form.get('valid_until')
        try:
            if not code.code or discount_type not in ('percentage', 'fixed') or code.value <= 0:
                raise ValueError('بيانات كود الخصم غير صالحة.')
            if discount_type == 'percentage' and code.value > 100:
                raise ValueError('نسبة الخصم لا يمكن أن تتجاوز 100%.')
            code.city_id, code.neighborhood_id = _parse_discount_location()
            code.valid_until = datetime.strptime(valid_until_str, '%Y-%m-%d')
        except (ValueError, TypeError) as exc:
            flash(str(exc), 'error')
            return render_template('admin/edit_discount_code.html', code=code, cities=cities, neighborhoods=neighborhoods)
        
        usage_limit = request.form.get('usage_limit')
        code.usage_limit = int(usage_limit) if usage_limit else None
        
        max_uses_per_customer = request.form.get('max_uses_per_customer')
        code.max_uses_per_customer = int(max_uses_per_customer) if max_uses_per_customer else 1
        
        code.is_active = 'is_active' in request.form
        
        try:
            db.session.commit()
            flash('تم تحديث كود الخصم بنجاح', 'success')
            return redirect(url_for('admin.discount_codes'))
        except:
            db.session.rollback()
            flash('حدث خطأ أثناء تحديث الكود.', 'error')
            
    return render_template('admin/edit_discount_code.html', code=code, cities=cities, neighborhoods=neighborhoods)

@bp.route('/discount_codes/delete/<int:id>', methods=['POST'])
def delete_discount_code(id):
    code = _get_scoped_discount_or_404(id)
    CheckoutSession.query.filter_by(recovery_discount_code_id=code.id).update(
        {'recovery_discount_code_id': None}, synchronize_session=False
    )
    db.session.delete(code)
    db.session.commit()
    flash('تم حذف كود الخصم بنجاح', 'success')
    return redirect(url_for('admin.discount_codes'))

@bp.route('/discount_codes/stats/<int:id>')
def discount_code_stats(id):
    code = _get_scoped_discount_or_404(id)
    bookings = Booking.query.filter_by(discount_code_id=id).all()
    total_savings = sum((b.service.price or 0) * ((code.value or 0) / 100) if code.discount_type == 'percentage' else (code.value or 0) for b in bookings if b.service)
    
    return render_template('admin/discount_code_stats.html', code=code, bookings=bookings, total_savings=total_savings)

# --- Admin Management ---
@bp.route('/admins')
def admins():
    page = request.args.get('page', 1, type=int)
    pagination = User.query.filter_by(role='admin').paginate(page=page, per_page=50, error_out=False)
    admins_list = pagination.items
    return render_template('admin/admins.html', admins=admins_list, pagination=pagination)

@bp.route('/admins/add', methods=['GET', 'POST'])
def add_admin():
    form = AdminUserForm()
    if form.validate_on_submit():
        # Check if username or email exists
        if User.query.filter_by(username=form.username.data).first():
            flash('اسم المستخدم موجود مسبقاً', 'error')
            return render_template('admin/admin_form.html', form=form, title='إضافة مسؤول')
        
        if User.query.filter_by(email=form.email.data).first():
            flash('البريد الإلكتروني مسجل مسبقاً لمستخدم آخر', 'error')
            return render_template('admin/admin_form.html', form=form, title='إضافة مسؤول')
        
        user = User(
            username=form.username.data,
            email=form.email.data,
            role='admin'
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('تم إضافة المسؤول بنجاح', 'success')
        return redirect(url_for('admin.admins'))
    return render_template('admin/admin_form.html', form=form, title='إضافة مسؤول')

@bp.route('/admins/edit/<int:id>', methods=['GET', 'POST'])
def edit_admin(id):
    admin = User.query.get_or_404(id)
    form = AdminUserForm(obj=admin)
    if form.validate_on_submit():
        admin.username = form.username.data
        admin.email = form.email.data
        if form.password.data:
            admin.set_password(form.password.data)
        db.session.commit()
        flash('تم تعديل بيانات المسؤول', 'success')
        return redirect(url_for('admin.admins'))
    return render_template('admin/admin_form.html', form=form, title='تعديل مسؤول', admin=admin)

@bp.route('/admins/delete/<int:id>', methods=['POST'])
def delete_admin(id):
    if id == current_user.id:
        flash('لا يمكنك حذف حسابك الحالي', 'error')
        return redirect(url_for('admin.admins'))
        
    admin = User.query.get_or_404(id)
    db.session.delete(admin)
    db.session.commit()
    flash('تم حذف المسؤول', 'success')
    return redirect(url_for('admin.admins'))


# ===== Gift Orders Management =====

@bp.route('/gift-orders')
def gift_orders():
    """List all gift orders with tabs for status"""
    from app.models import GiftOrder
    
    status_filter = request.args.get('status', 'pending')
    
    # Get base query
    base_query = GiftOrder.query
    
    # Filter for supervisors - only show gifts for their neighborhoods
    if current_user.role == 'supervisor':
        supervisor_neighborhood_ids = []
        if current_user.supervisor_neighborhoods:
            supervisor_neighborhood_ids.extend([n.id for n in current_user.supervisor_neighborhoods])
        
        if current_user.supervisor_cities:
            for city in current_user.supervisor_cities:
                supervisor_neighborhood_ids.extend([n.id for n in city.neighborhoods])
        
        if supervisor_neighborhood_ids:
            base_query = base_query.filter(GiftOrder.neighborhood_id.in_(supervisor_neighborhood_ids))
        else:
            base_query = base_query.filter_by(id=-1)  # Empty result
    
    page = request.args.get('page', 1, type=int)
    pagination = base_query.filter_by(status=status_filter).order_by(GiftOrder.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    orders = pagination.items
    
    pending_count = base_query.filter_by(status='pending').count()
    accepted_count = base_query.filter_by(status='accepted').count()
    rejected_count = base_query.filter_by(status='rejected').count()
    
    return render_template('admin/gift_orders.html',
                         orders=orders,
                         pending_count=pending_count,
                         accepted_count=accepted_count,
                         rejected_count=rejected_count,
                         status_filter=status_filter,
                         pagination=pagination)


@bp.route('/gift-orders/<int:id>/accept', methods=['POST'])
def accept_gift_order(id):
    """Accept a gift order"""
    from app.models import GiftOrder
    
    gift_order = GiftOrder.query.get_or_404(id)
    gift_order.status = 'accepted'
    db.session.commit()
    
    flash('تم قبول طلب الهدية', 'success')
    return redirect(url_for('admin.gift_orders'))


@bp.route('/gift-orders/<int:id>/reject', methods=['POST'])
def reject_gift_order(id):
    """Reject a gift order"""
    from app.models import GiftOrder
    
    gift_order = GiftOrder.query.get_or_404(id)
    gift_order.status = 'rejected'
    db.session.commit()
    
    flash('تم رفض طلب الهدية', 'warning')
    return redirect(url_for('admin.gift_orders'))


# --- Announcement Management ---
@bp.route('/announcements')
def announcements():
    """List all announcements"""
    page = request.args.get('page', 1, type=int)
    pagination = Announcement.query.order_by(Announcement.order.asc(), Announcement.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    announcements_list = pagination.items
    return render_template('admin/announcements.html', announcements=announcements_list, pagination=pagination)


@bp.route('/announcements/add', methods=['GET', 'POST'])
def add_announcement():
    """Add a new announcement"""
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        link_url = request.form.get('link_url')
        order = int(request.form.get('order', 0))
        is_active = 'is_active' in request.form
        
        # Handle image upload
        image_url = ''
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                from app.utils.file_handling import allowed_file
                if not allowed_file(file.filename):
                    flash('نوع الملف غير مدعوم. يرجى رفع صورة (png, jpg, jpeg, gif, webp)', 'error')
                    return redirect(url_for('admin.announcements'))
                
                filename = secure_filename(file.filename)
                # Create unique filename
                unique_filename = f"announcement_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
                upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'announcements')
                os.makedirs(upload_folder, exist_ok=True)
                file.save(os.path.join(upload_folder, unique_filename))
                image_url = f"/static/uploads/announcements/{unique_filename}"
        
        announcement = Announcement(
            title=title,
            description=description,
            image_url=image_url,
            link_url=link_url,
            order=order,
            is_active=is_active
        )
        db.session.add(announcement)
        db.session.commit()
        
        flash('تم إضافة الإعلان بنجاح', 'success')
        return redirect(url_for('admin.announcements'))
    
    return render_template('admin/announcement_form.html', announcement=None)


@bp.route('/announcements/edit/<int:id>', methods=['GET', 'POST'])
def edit_announcement(id):
    """Edit an announcement"""
    announcement = Announcement.query.get_or_404(id)
    
    if request.method == 'POST':
        announcement.title = request.form.get('title')
        announcement.description = request.form.get('description')
        announcement.link_url = request.form.get('link_url')
        announcement.order = int(request.form.get('order', 0))
        announcement.is_active = 'is_active' in request.form
        
        # Handle image upload
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                from app.utils.file_handling import allowed_file
                if not allowed_file(file.filename):
                    flash('نوع الملف غير مدعوم. يرجى رفع صورة (png, jpg, jpeg, gif, webp)', 'error')
                    return redirect(url_for('admin.announcements'))

                filename = secure_filename(file.filename)
                unique_filename = f"announcement_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
                upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'announcements')
                os.makedirs(upload_folder, exist_ok=True)
                file.save(os.path.join(upload_folder, unique_filename))
                announcement.image_url = f"/static/uploads/announcements/{unique_filename}"
        
        db.session.commit()
        flash('تم تحديث الإعلان بنجاح', 'success')
        return redirect(url_for('admin.announcements'))
    
    return render_template('admin/announcement_form.html', announcement=announcement)


@bp.route('/announcements/delete/<int:id>', methods=['POST'])
def delete_announcement(id):
    """Delete an announcement"""
    announcement = Announcement.query.get_or_404(id)
    db.session.delete(announcement)
    db.session.commit()
    flash('تم حذف الإعلان', 'success')
    return redirect(url_for('admin.announcements'))


@bp.route('/announcements/toggle/<int:id>', methods=['POST'])
def toggle_announcement(id):
    """Toggle announcement active status"""
    announcement = Announcement.query.get_or_404(id)
    announcement.is_active = not announcement.is_active
    db.session.commit()
    status = 'مفعل' if announcement.is_active else 'معطل'
    flash(f'تم تحديث حالة الإعلان إلى {status}', 'success')
    return redirect(url_for('admin.announcements'))


# --- Employee Location Tracking ---
@bp.route('/employee-tracking')
def employee_tracking():
    """Admin page to track employee locations on a map"""
    employees = _scoped_employee_query(include_break=True).all()
    return render_template('admin/employee_tracking.html', employees=employees)


@bp.route('/api/employee-locations')
def get_employee_locations():
    """API endpoint to get all active employee locations with enhanced data"""
    from datetime import datetime, timedelta
    
    locations_query = EmployeeLocation.query.filter_by(is_tracking=True)
    neighborhood_ids = _supervisor_neighborhood_ids()
    allowed_employee_ids = None
    if neighborhood_ids is not None:
        allowed_employee_ids = [emp.id for emp in _scoped_employee_query(include_break=True).all()]
        locations_query = locations_query.filter(EmployeeLocation.employee_id.in_(allowed_employee_ids) if allowed_employee_ids else EmployeeLocation.employee_id == -1)
    locations = locations_query.all()
    
    result = []
    for loc in locations:
        # Check if location was updated in the last 10 minutes
        is_recent = loc.updated_at > datetime.utcnow() - timedelta(minutes=10)
        seconds_since_update = (datetime.utcnow() - loc.updated_at).total_seconds()
        
        # Count active bookings today
        from app.utils.timezone import get_saudi_date
        today = get_saudi_date()
        active_bookings_count = Booking.query.filter(
            Booking.employee_id == loc.employee_id,
            Booking.date == today,
            Booking.status.in_(['assigned', 'en_route', 'arrived', 'in_progress'])
        ).count()
        
        # Check for active booking
        active_booking = Booking.query.filter(
            Booking.employee_id == loc.employee_id,
            Booking.status.in_(['assigned', 'en_route', 'arrived', 'in_progress'])
        ).first()
        
        booking_info = None
        if active_booking:
            booking_info = {
                'id': active_booking.id,
                'status': active_booking.status,
                'status_ar': {
                    'assigned': 'تم التعيين',
                    'en_route': 'في الطريق',
                    'arrived': 'وصل',
                    'in_progress': 'جاري العمل'
                }.get(active_booking.status, active_booking.status)
            }
        
        # Get employee neighborhoods
        employee = loc.employee
        neighborhood_name = ''
        if employee and employee.neighborhoods:
            neighborhood_name = ', '.join([n.name_ar for n in employee.neighborhoods[:3]])
            
        # Convert time to Saudi Time (UTC+3)
        saudi_time = loc.updated_at + timedelta(hours=3)
        
        result.append({
            'employee_id': loc.employee_id,
            'employee_name': employee.username if employee else 'Unknown',
            'name': employee.username if employee else 'Unknown',
            'phone': employee.phone if employee else '',
            'latitude': loc.latitude,
            'longitude': loc.longitude,
            'lat': loc.latitude,
            'lng': loc.longitude,
            'accuracy': loc.accuracy,
            'updated_at': saudi_time.strftime('%I:%M:%S %p'),
            'seconds_since_update': seconds_since_update,
            'is_recent': is_recent,
            'neighborhood_name': neighborhood_name,
            'active_bookings_count': active_bookings_count,
            'booking': booking_info
        })
    
    return jsonify(result)


# --- City-Based Pricing: Services ---
@bp.route('/api/city-service-prices/<int:service_id>')
def get_city_service_prices(service_id):
    """Get all city & size prices for a service"""
    from app.models import City, VehicleSize
    prices = CityServicePrice.query.filter_by(service_id=service_id).all()
    return jsonify([{
        'id': p.id,
        'city_id': p.city_id,
        'city_name': City.query.get(p.city_id).name_ar if City.query.get(p.city_id) else '',
        'size_id': p.vehicle_size_id,
        'size_name': VehicleSize.query.get(p.vehicle_size_id).name_ar if VehicleSize.query.get(p.vehicle_size_id) else '',
        'price': p.price,
        'is_active': p.is_active
    } for p in prices])


@bp.route('/services/assign-city-size', methods=['POST'])
def assign_service_to_city_size():
    """Assign a service to a city and vehicle size with a specific price"""
    service_id = request.form.get('service_id', type=int)
    city_id = request.form.get('city_id', type=int)
    vehicle_size_id = request.form.get('vehicle_size_id', type=int)
    price = request.form.get('price', type=float)
    
    if not all([service_id, city_id, vehicle_size_id, price is not None]):
        flash('جميع الحقول مطلوبة', 'error')
        return redirect(url_for('admin.services'))
    
    # Update or Create
    existing = CityServicePrice.query.filter_by(
        city_id=city_id, 
        service_id=service_id, 
        vehicle_size_id=vehicle_size_id
    ).first()
    
    if existing:
        existing.price = price
        flash('تم تحديث السعر المخصص بنجاح', 'success')
    else:
        csp = CityServicePrice(
            city_id=city_id, 
            service_id=service_id, 
            vehicle_size_id=vehicle_size_id, 
            price=price, 
            is_active=True
        )
        db.session.add(csp)
        flash('تم إسناد السعر المخصص بنجاح', 'success')
    
    db.session.commit()
    return redirect(url_for('admin.services'))


@bp.route('/services/update-city-price', methods=['POST'])
def update_service_city_price():
    """Update city price for a service"""
    price_id = request.form.get('price_id', type=int)
    new_price = request.form.get('price', type=float)
    
    csp = CityServicePrice.query.get_or_404(price_id)
    csp.price = new_price
    db.session.commit()
    flash('تم تحديث السعر بنجاح', 'success')
    return redirect(url_for('admin.services'))


@bp.route('/services/remove-city-price/<int:price_id>', methods=['POST'])
def remove_service_city_price(price_id):
    """Remove a city price assignment for a service"""
    csp = CityServicePrice.query.get_or_404(price_id)
    db.session.delete(csp)
    db.session.commit()
    flash('تم إزالة الخدمة من المدينة', 'success')
    return redirect(url_for('admin.services'))


@bp.route('/services/duplicate/<int:id>', methods=['POST'])
def duplicate_service(id):
    """Duplicate a service"""
    service = Service.query.get_or_404(id)
    new_service = Service(
        name_ar=service.name_ar + ' - نسخة',
        name_en=service.name_en + ' - Copy',
        price=service.price,
        duration=service.duration,
        description=service.description,
        includes_free_wash=service.includes_free_wash,
        awards_loyalty_point=service.awards_loyalty_point,
        is_active=service.is_active
    )
    db.session.add(new_service)
    db.session.commit()
    flash(f'تم نسخ الخدمة بنجاح (#{new_service.id}). يمكنك الآن إسنادها لمدينة.', 'success')
    return jsonify({'status': 'ok', 'new_service_id': new_service.id})


# --- City-Based Pricing: Products ---
@bp.route('/api/city-product-prices/<int:product_id>')
def get_city_product_prices(product_id):
    """Get all city prices for a product"""
    prices = CityProductPrice.query.filter_by(product_id=product_id).all()
    return jsonify([{
        'id': p.id,
        'city_id': p.city_id,
        'city_name': City.query.get(p.city_id).name_ar if City.query.get(p.city_id) else '',
        'price': p.price,
        'is_active': p.is_active
    } for p in prices])


@bp.route('/products/assign-city', methods=['POST'])
def assign_product_to_city():
    """Assign a product to a city with a specific price"""
    product_id = request.form.get('product_id', type=int)
    city_id = request.form.get('city_id', type=int)
    price = request.form.get('price', type=float)
    
    if not all([product_id, city_id, price is not None]):
        flash('بيانات غير مكتملة', 'error')
        return redirect(url_for('admin.products'))
    
    existing = CityProductPrice.query.filter_by(city_id=city_id, product_id=product_id).first()
    if existing:
        flash('المنتج مسند لهذه المدينة بالفعل', 'error')
        return redirect(url_for('admin.products'))
    
    cpp = CityProductPrice(city_id=city_id, product_id=product_id, price=price, is_active=True)
    db.session.add(cpp)
    db.session.commit()
    flash('تم إسناد المنتج للمدينة بنجاح', 'success')
    return redirect(url_for('admin.products'))


@bp.route('/products/update-city-price', methods=['POST'])
def update_product_city_price():
    """Update city price for a product"""
    price_id = request.form.get('price_id', type=int)
    new_price = request.form.get('price', type=float)
    
    cpp = CityProductPrice.query.get_or_404(price_id)
    cpp.price = new_price
    db.session.commit()
    flash('تم تحديث السعر بنجاح', 'success')
    return redirect(url_for('admin.products'))


@bp.route('/products/remove-city-price/<int:price_id>', methods=['POST'])
def remove_product_city_price(price_id):
    """Remove a city price assignment for a product"""
    cpp = CityProductPrice.query.get_or_404(price_id)
    db.session.delete(cpp)
    db.session.commit()
    flash('تم إزالة المنتج من المدينة', 'success')
    return redirect(url_for('admin.products'))


@bp.route('/products/duplicate/<int:id>', methods=['POST'])
def duplicate_product(id):
    """Duplicate a product"""
    product = Product.query.get_or_404(id)
    new_product = Product(
        name_ar=product.name_ar + ' - نسخة',
        name_en=product.name_en + ' - Copy',
        price=product.price,
        image_url=product.image_url,
        stock_quantity=0,
        is_active=product.is_active
    )
    db.session.add(new_product)
    db.session.commit()
    flash(f'تم نسخ المنتج بنجاح (#{new_product.id}). يمكنك الآن إسناده لمدينة.', 'success')
    return jsonify({'status': 'ok', 'new_product_id': new_product.id})


# ============== Referral Tracking ==============

@bp.route('/referrals')
def referral_tracking():
    """Admin referral tracking page"""
    from app.models import ReferralRecord, User
    
    page = request.args.get('page', 1, type=int)
    pagination = ReferralRecord.query.order_by(ReferralRecord.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    all_records = pagination.items
    
    total_referrals = pagination.total
    completed_washes = ReferralRecord.query.filter_by(first_wash_completed=True).count()
    pending_washes = total_referrals - completed_washes
    conversion_rate = round(completed_washes / total_referrals * 100, 1) if total_referrals > 0 else 0
    
    # Top referrers - users with most referrals
    from sqlalchemy import func
    top_referrer_data = db.session.query(
        ReferralRecord.referrer_id,
        func.count(ReferralRecord.id).label('total'),
        func.sum(db.case((ReferralRecord.first_wash_completed == True, 1), else_=0)).label('completed')
    ).group_by(ReferralRecord.referrer_id).order_by(func.count(ReferralRecord.id).desc()).limit(20).all()
    
    top_referrers = []
    for referrer_id, total, completed in top_referrer_data:
        user = User.query.get(referrer_id)
        if user:
            top_referrers.append({
                'user': user,
                'total': total,
                'completed': completed or 0
            })
    
    return render_template('admin/referrals.html',
                         all_records=all_records,
                         total_referrals=total_referrals,
                         completed_washes=completed_washes,
                         pending_washes=pending_washes,
                         conversion_rate=conversion_rate,
                         top_referrers=top_referrers,
                         pagination=pagination)


# ============== Influencer Codes Management ==============

@bp.route('/influencer-codes')
def influencer_codes():
    """Admin influencer codes management page"""
    from app.models import DiscountCode
    page = request.args.get('page', 1, type=int)
    query = DiscountCode.query.filter_by(is_influencer=True)
    pagination = query.order_by(DiscountCode.created_at.desc() if hasattr(DiscountCode, 'created_at') else DiscountCode.valid_from.desc()).paginate(page=page, per_page=50, error_out=False)
    codes = pagination.items
    # Calculate total usage across ALL influencer codes
    total_usage = db.session.query(func.sum(DiscountCode.used_count)).filter_by(is_influencer=True).scalar() or 0
    total_codes = pagination.total
    return render_template('admin/influencer_codes.html',
                         total_usage=total_usage,
                         total_codes=total_codes,
                         codes=codes,
                         pagination=pagination)


@bp.route('/influencer-codes/add', methods=['POST'])
def add_influencer_code():
    """Add a new influencer code"""
    from app.models import DiscountCode
    from datetime import datetime
    
    code_value = request.form.get('code', '').strip().upper()
    name = request.form.get('name', '').strip()
    discount = request.form.get('discount', type=float)
    
    if not code_value or not name or discount is None:
        flash('يرجى تعبئة جميع الحقول', 'error')
        return redirect(url_for('admin.influencer_codes'))
    
    existing = DiscountCode.query.filter_by(code=code_value).first()
    if existing:
        flash('هذا الكود موجود بالفعل', 'error')
        return redirect(url_for('admin.influencer_codes'))
    
    new_code = DiscountCode(
        code=code_value,
        discount_type='percentage',
        value=discount,
        valid_until=datetime(2099, 12, 31),  # effectively no expiry
        is_influencer=True,
        influencer_name=name
    )
    db.session.add(new_code)
    db.session.commit()
    flash(f'تم إضافة كود المؤثر "{code_value}" بنجاح', 'success')
    return redirect(url_for('admin.influencer_codes'))


@bp.route('/influencer-codes/<int:id>/toggle', methods=['POST'])
def toggle_influencer_code(id):
    """Toggle influencer code active/inactive"""
    from app.models import DiscountCode
    code = DiscountCode.query.get_or_404(id)
    if not code.is_influencer:
        return redirect(url_for('admin.influencer_codes'))
        
    code.is_active = not code.is_active
    db.session.commit()
    status = 'مفعّل' if code.is_active else 'معطّل'
    flash(f'تم تحديث حالة الكود "{code.code}" إلى {status}', 'success')
    return redirect(url_for('admin.influencer_codes'))


@bp.route('/influencer-codes/edit/<int:id>', methods=['GET', 'POST'])
def edit_influencer_code(id):
    """Edit an influencer code"""
    code = DiscountCode.query.get_or_404(id)
    
    if not code.is_influencer:
        return redirect(url_for('admin.influencer_codes'))
        
    if request.method == 'POST':
        code.code = request.form.get('code', '').strip().upper()
        code.influencer_name = request.form.get('influencer_name', '').strip()
        code.value = float(request.form.get('value', 0))
        code.discount_type = request.form.get('discount_type', 'percentage').lower()
        
        # Valid until
        valid_until_str = request.form.get('valid_until')
        if valid_until_str:
            code.valid_until = datetime.strptime(valid_until_str, '%Y-%m-%d')
            
        usage_limit = request.form.get('usage_limit')
        code.usage_limit = int(usage_limit) if usage_limit else None
        
        max_uses_per_customer = request.form.get('max_uses_per_customer')
        code.max_uses_per_customer = int(max_uses_per_customer) if max_uses_per_customer else 1
        
        code.is_active = 'is_active' in request.form
        
        try:
            db.session.commit()
            flash('تم تحديث بيانات كود المؤثر بنجاح', 'success')
            return redirect(url_for('admin.influencer_codes'))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ أثناء التحديث', 'error')
            
    return render_template('admin/edit_influencer_code.html', code=code)



@bp.route('/influencer-codes/<int:id>/delete', methods=['POST'])
def delete_influencer_code(id):
    """Delete an influencer code"""
    from app.models import DiscountCode
    code = DiscountCode.query.get_or_404(id)
    if not code.is_influencer:
        return redirect(url_for('admin.influencer_codes'))
        
    code_name = code.code
    db.session.delete(code)
    db.session.commit()
    flash(f'تم حذف كود المؤثر "{code_name}" بنجاح', 'success')
    return redirect(url_for('admin.influencer_codes'))

